import re
import os
import sys
from openpyxl import load_workbook
import configparser
from stop_cltk import STOPS_LIST

os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')

file_list = config['paths']['file_list']
af = config['paths']['annotated']
lf = config['paths']['output']
filterStopWords=config['params']['filter_stop_words']

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws['A1:U1']
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws['A2:U803']

os.system("clear && printf '\e[3J'")

for record in files:
	file = '%s/%s'%(af,record[h['Tokenized files']].value)
	try:
		finalTxt = open(file, 'r').read().replace("\n", "").replace(" ", "")
		finalTxt = re.sub('.*?<body>(.*?)</body>.*', r'\1', finalTxt)
		finalTxt = re.sub('<punct.*?/>', '', finalTxt)
		finalTxt = re.sub('<sentenceid="(.*?)"location="(.*?)">', r'\1 (\2): ', finalTxt)
		finalTxt = re.sub('<word.*?entry="(.*?)".*?</word>', r'\1 ', finalTxt)
		finalTxt = re.sub('</sentence>', '\n', finalTxt)
		if filterStopWords == "True":
			for stop in STOPS_LIST:
				finalTxt = finalTxt.replace(' %s '%stop, ' ')
		open('%s/%s'%(lf,record[h['Tokenized files']].value.replace('xml', 'txt')),'w').write(finalTxt)
		print(record[h['Tokenized files']].value, 'done')
	except:
		pass