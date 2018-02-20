from openpyxl import load_workbook
import os
import re
import configparser
from grkLemmata import greekLemmata

os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')

resources = config['paths']['file_list']
lf = config['paths']['output']

wb = load_workbook('%s/word_senses.xlsx'%resources)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
data = ws['A2:L8']

print('-- File list --\n')
fileListDict = {}
w = 0
for file in os.listdir(lf):
	if file.endswith(".txt"):
		w+=1
		print(w,':',file)
		fileListDict[w]=file
file = int(input('\nInsert file number: '))
while True :
	try:
		file_selected=fileListDict[file]
		file_contents=open("%s/%s"%(lf,file_selected),'r').read()
		print('File selected:',file_selected,'\n')
	except:
		print('Wrong input')
		file = int(input('Insert file number: '))		
	else:
		break

print('-- Words in database --\n')
words = {}
for record in data:
	words[record[h['TERM']].value]=record[h['TERM ID']].value

for idx,(term,t_id) in enumerate(words.items()):
	print('%s: %s (%s)'%(str(idx+1),term,t_id))

print()

while True:
	term = int(input('Insert word number: '))
	if term > 0 and term <= len(words):
		break
		
for idx,(form,t_id) in enumerate(words.items()):
	if idx == term-1:
		current_id = t_id
		word_selected='%s (%s)'%(form,t_id)
		print('Word selected:',word_selected)
		
		
things_to_check = re.findall('p\(w\|s\).*', file_contents)

output={}
for thing in things_to_check:
	current_set=re.search('p\(w\|s\) = sum_t P\(w\|t,s\)  (T=-?\d+,K=\d+)',thing).group(1)
	output.setdefault(current_set,{})
	for x in re.findall('[:;]\s+(.*? \[(.*?)\] \(.*?\)) ',thing):
		current_word_list = ''
		for record in data:
			if record[h['TERM ID']].value == current_id:
				current_word_list = record[h['CONTEXT WORDS IDS']].value
				if current_word_list != None:
					current_word_list=current_word_list.split()
					output[current_set].setdefault(record[h['SENSE']].value,[])
					if x[1] in current_word_list:
					#word found in word list
						output[current_set][record[h['SENSE']].value].append(x[0])
#print(output)
output_string =''
indent = False
for key,value in output.items():
	current_set=key
	output_string='%s%s:'%(output_string,current_set)
	indent = False
	for senses,words in value.items():
		if len(words) > 0:
			if indent == True: output_string += "\t"
			current_sense = senses
			current_words='; '.join(words)
			output_string='%s\t%s\t%s\n'%(output_string,current_sense, current_words)
			indent=True
	output_string+="\n"
print('\n',20*'#','\n', sep='')
print(output_string.strip())
print('\n',20*'#', sep='')
filename = 'Senses of {word} in {file}'.format(file=file_selected,word=word_selected)
open("%s/%s.txt"%(lf,filename),'w').write(output_string)
			