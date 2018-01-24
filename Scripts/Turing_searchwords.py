import os
import re
import os
import sys
import lxml.etree as document
from openpyxl import load_workbook
from tlgIndex import tlgIndexes
from beta2utf import convertBeta
from utf2beta import convertUTF
import configparser

os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')
file_list = config['paths']['file_list']
spreadsheet_output = config['paths']['spreadsheet_output']
af = config['paths']['annotated']

os.system("clear && printf '\e[3J'")

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws['A1:U1']
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws['A2:U803']

wb2 = load_workbook('%s/Word frequencies in corpus (preliminary 3).xlsx'%spreadsheet_output)
ws2 = wb2.active
entries = ws2['B2:V2']
authors = {}
totalThing = len(files)*len(entries[0])
currentIter = 0
for record in files:
	file = '%s/%s'%(af,record[6].value)
	parse = document.parse(file)
	authorId = parse.xpath('//tlgAuthor')[0].text
	#print(authorId)
	if len(authorId)>4: authorId = authorId[-4:]
	try:
		authors[authorId]
	except:
		authors[authorId] = {}
	for entry in entries[0]:
		currentIter += 1
		lexeme = entry.value
		count = 0
		count = len(parse.xpath('//word[lemma/@id="%s"]'%lexeme))
		#print('%s: %d'%(entry.value, count))
		try:
			authors[authorId][lexeme]
		except:
			authors[authorId][lexeme] = 0
		authors[authorId][lexeme] += count
		progPercent = str(round(currentIter*100/totalThing, 2))
		sys.stdout.write("\r\033[KProgress: %s%%"%(progPercent))
		sys.stdout.flush()
for row,(author,lexeme) in enumerate(authors.items()):
	author = author.strip()
	ws2['A%s'%str(row+3)].value = tlgIndexes[author]
	for col,(entry,count) in enumerate(lexeme.items()):
		#print('%s: %d'%(value1, value2))
		ws2['%s%d'%(chr(ord('B')+col),row+3)].value = count
print('\nAll done!')
wb2.save('%s/Word frequencies in corpus (preliminary 3).xlsx'%spreadsheet_output)
	
