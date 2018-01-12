import re
import os
import sys
import lxml.etree as document
from utf2beta import convertUTF
from beta2utf import convertBeta
from grkFrm import greekForms 
from grkLemmata import greekLemmata
import time
from openpyxl import load_workbook
import configparser

os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')
file_list = config['paths']['file_list']
tf = config['paths']['tokenized']
af = config['paths']['annotated']
logf = config['paths']['logs']

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
files = ws['A2:Q803']

os.system("clear && printf '\e[3J'")
attrName=['lemma', 'translation', 'morph']
unknownlist = open('%s/unknown_list.txt'%logf, 'a')
unknownlist_proper = open('%s/unknown_list_proper.txt'%logf, 'a')
unknownlist_single = open('%s/unknown_list_single.txt'%logf, 'a')

#####FUNCTIONS#####
def cleanWords(word):
	word = re.sub('[^a-z\(\)\\\/\*\|=\+\'0-9\"]*', '', word)

	#turn graves into acutes, lowercase betacode, remove clitic accent, escape diacritics
	wordForm = word.replace("2", "").replace("3", "")
	word = word.replace('\\', '/').lower()
	word = re.sub(r"\*([aehiouw])([\(\)/=]+)",r"*\2\1", word) #parole registrate maiuscole: *DIACRlettere; parole minuscole: *parola normale, quindi * + e)/ etc.
	word = re.sub(r"([\(\)/=]+)(\*)([aehiouw])",r"\2\1\3", word)
	word = re.sub(r"\|([/\(\)])",r"\g<1>|", word)
	howmanyAccents = len(re.findall("[/=]", word))
	if howmanyAccents > 1:
		word=word[::-1].replace("/", "", 1)[::-1]
	return word
	
def lookup(word):
	word = cleanWords(word)
	try:
		entry=greekForms[word]
	#no match			
	except KeyError:
		word = word.replace("+", "")
		word = re.sub(r"\*([\(\)/=]+)([aehiouw])",r"*\2\1", word)
		word = word.replace("\*", "")
		try:
			entry=greekForms[word]
		except KeyError:
			word = word.replace("'", "")
			try:
				entry=greekForms[word]
			except KeyError:
				word = re.sub("\)$", "'",word)
				try:
					entry=greekForms[word]
				except KeyError:
					word = re.sub("'$", "",word)
					try:
						entry=greekForms[word]
					except KeyError:
						return {'is_unknown':True}
					else:
						return {'is_unknown':False, 'entry':entry}
				else:
					return {'is_unknown':False, 'entry':entry}
			else:
				return {'is_unknown':False, 'entry':entry}
		else:
			return {'is_unknown':False, 'entry':entry}
	else:
		return {'is_unknown':False, 'entry':entry}

def process(lemmata, word):
	#each wordform corresponds to a list of lemmata
	global wml
	if len(lemmata)>1: wml+=1
	for lemma in lemmata:
	#each lemma is a dictionary; 'lemma' = entry, 'analysis' is a list	
		newLemma = document.SubElement(word, "lemma")
		newLemma.set('id',lemma['lemma'])
		newLemma.set('entry',greekLemmata[lemma['lemma']]['lemma'])
		newLemma.set('POS',greekLemmata[lemma['lemma']]['pos'])
		for analysis in lemma['analysis']:
			newAnalysis = document.SubElement(newLemma, "analysis")
			newAnalysis.set('morph', analysis)
			del newAnalysis
		del newLemma
		
###################

for record in files:
	file = '%s/%s'%(tf,record[6].value)
	unknowncount = 0
	unknowncount_proper = 0
	unknowncount_single = 0
	wml = 0 #words with multiple lemmata
	print('%s - %s'%(record[3].value,record[4].value))
	progPercent=0
	fileName = record[6].value
	if os.path.exists('%s/%s'%(af, fileName)):
		continue
	parse = document.parse(file)
	words = parse.xpath('//word')
	total = len(words)
	for idx, word in enumerate(words):
		progPercent = str(int(idx*100/total))
		sys.stdout.write("\r\033[K\tProgress: %s%%"%(progPercent))
		sys.stdout.flush()
		#get word form
		form = word.get("form")
		isunknown = False
		#lookup in dictionary
		if lookup(form)['is_unknown'] == True:
		#####test lacunae (merge with following words)#####
			try:
				next_word = words[idx+1]
			except:
				isunknown = True
			else:
				next_form = next_word.get("form")
				iter = 0
				while True:
					iter += 1
					previous_word = words[idx-iter]
					if previous_word.get('toremove') == None:
						break
				previous_form = previous_word.get("form")
				merged_form = ''
				if word.get('lacuna') != None:
					merged_form = '%s%s'%(form, next_form)
					if lookup(merged_form)['is_unknown'] == True:
						try:
							next_next_word = words[idx+2]
						except:
							isunknown = True
						else:
							next_next_form = next_next_word.get("form")
							if next_next_word.get('lacuna') != None:
								merged_form='%s%s'%(merged_form, next_next_form)
								if lookup(merged_form)['is_unknown'] == True:
									try:
										next_next_next_word = words[idx+3]
									except:
										isunknown = True
									else:
										next_next_next_form = next_next_next_word.get("form")
										merged_form = '%s%s'%(merged_form, next_next_next_form)
										if lookup(merged_form)['is_unknown'] == True:
											isunknown = True
										else:
											word.set('form', merged_form)
											process(lookup(merged_form)['entry'], word)
											next_word.set('toremove', 'true')
											next_next_word.set('toremove', 'true')
											next_next_next_word.set('toremove', 'true')
								else:
									word.set('form', merged_form)
									process(lookup(merged_form)['entry'], word)
									next_word.set('toremove', 'true')
									next_next_word.set('toremove', 'true')
							else:
								isunknown = True
					else:
						word.set('form', merged_form)
						process(lookup(merged_form)['entry'], word)
						next_word.set('toremove', 'true')
				else:
					if next_word.get('lacuna') != None:
						merged_form = '%s%s'%(form, next_form)
						if lookup(merged_form)['is_unknown'] == True:
							try:
								next_next_word = words[idx+2]
							except:
								isunknown = True
							else:
								next_next_form = next_next_word.get("form")
								merged_form='%s%s'%(merged_form, next_next_form)
								if lookup(merged_form)['is_unknown'] == True:
									try:
										next_next_next_word = words[idx+3]
									except:
										isunknown = True
									else:
										next_next_next_form = next_next_next_word.get("form")
										merged_form = '%s%s'%(merged_form, next_next_next_form)
										if lookup(merged_form)['is_unknown'] == True:
											isunknown = True
										else:
											word.set('form', merged_form)
											process(lookup(merged_form)['entry'], word)
											next_word.set('toremove', 'true')
											next_next_word.set('toremove', 'true')
											next_next_next_word.set('toremove', 'true')
								else:
									word.set('form', merged_form)
									word.set('lacuna', 'true')
									process(lookup(merged_form)['entry'], word)
									next_word.set('toremove', 'true')
									next_next_word.set('toremove', 'true')
						else:
							word.set('form', merged_form)
							word.set('lacuna', 'true')
							process(lookup(merged_form)['entry'], word)
							next_word.set('toremove', 'true')
					else:
						if form[-1] == '-':
							merged_form = '%s%s'%(form, next_form)
							if lookup(merged_form)['is_unknown'] == True:
								isunknown = True
							else:
								word.set('form', merged_form)
								process(lookup(merged_form)['entry'], word)
								next_word.set('toremove', 'true')
						else:
							isunknown = True
		#####end test lacunae (1)#####
		#####test lacunae (merge with previous word and next word)#####
			if isunknown == True and (word.get('lacuna') != None or previous_word.get('lacuna') != None):
				merged_form='%s%s'%(previous_form, form)
				if lookup(merged_form)['is_unknown'] == True:
					merged_form='%s%s'%(merged_form, next_form)
					if lookup(merged_form)['is_unknown'] == True:
						isunknown = True
					else:
						previous_word.set('form', merged_form)
						previous_word.set('lacuna', 'true')
						process(lookup(merged_form)['entry'], previous_word)
						word.set('toremove', 'true')
						next_word.set('toremove', 'true')
				else:
					previous_word.set('form', merged_form)
					previous_word.set('lacuna', 'true')
					process(lookup(merged_form)['entry'], previous_word)
					word.set('toremove', 'true')
		#####end test lacunae (2)#####

			if isunknown == True:
				document.SubElement(word, "lemma").set("id", "unknown")
		else:
			process(lookup(form)['entry'], word)

	#remove joined words
	toremove=parse.xpath('//word[@toremove]')
	for nTr in toremove:
		nTr.getparent().remove(nTr)
	
	#recalculate indexes
	for sentence in parse.xpath('//sentence'):
		words = sentence.xpath('./word')
		for idx, word in enumerate(words):
			word.set("id", str(idx+1))
			if word.xpath('./lemma')[0].get('id')=='unknown':	
				if word.get('form')[0] == '*':
					unknownlist_proper.write('%s : %s\n'%(word.get('form'), fileName))
					unknowncount_proper+=1
				elif len(word.get('form')) == 1:
					unknownlist_single.write('%s : %s\n'%(word.get('form'), fileName))
					unknowncount_single+=1
				else:
					unknownlist.write('%s : %s\n'%(word.get('form'), fileName))
					unknowncount+=1
		
	record[7].value = len(parse.xpath('//word'))
	record[8].value = unknowncount
	record[9].value = unknowncount_proper
	record[10].value = unknowncount_single
	record[15].value = wml
	parse.write('%s/%s'%(af, fileName), xml_declaration = True, encoding='UTF-8', pretty_print=True)
	wb.save('%s/file_list.xlsx'%file_list)
	print()
unknownlist_single.close()
unknownlist_proper.close()
unknownlist.close()