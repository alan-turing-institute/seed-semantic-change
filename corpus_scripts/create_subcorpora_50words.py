import re
import os
import sys
from openpyxl import load_workbook
import configparser
from grkLemmata import greekLemmata

os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))

config = configparser.ConfigParser()
config.read('config.ini')

file_list = config['paths']['file_list']
af = config['paths']['final_corpus']
dir = config['paths']['output']

wb = load_workbook('%s/50 words.xlsx'%dir)
ws = wb.active
words = ws['B1:B50']

file = open('%s/full_corpus_ids.txt'%dir, 'r')

#for each word, select sentences --> set, one item per sentence --> training
#count total
#random assignment (pop) of 20% to another set -- >test
#store sets as files in subfolder subcorpora/training and subcorpora/test
for word in words:
	sys.stdout.write("\r\033[KCreating subcorpora for: %s"%greekLemmata[str(word[0].value)]['lemma'])
	sys.stdout.flush()
	training = set()
	testing = set()
	for line in file:
		if str(word[0].value) in line.split('\t')[1].split():
			training.add(line)
	file.seek(0)
	for twenty_percent in range (0, int(len(training)*0.2)):
		testing.add(training.pop())
	open('%s/subcorpora/training/%s.txt'%(dir,str(word[0].value)), 'w').write(''.join([x for x in training]))
	open('%s/subcorpora/test/%s.txt'%(dir,str(word[0].value)), 'w').write(''.join([x for x in testing]))
print('All done! What a joy!')



