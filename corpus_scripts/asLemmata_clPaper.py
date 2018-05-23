import re
import os
import sys
from openpyxl import load_workbook
import configparser
from stop_cltk import STOPS_LIST_ID
from grkLemmata import greekLemmata
from lxml import etree as document
os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))

config = configparser.ConfigParser()
config.read('config.ini')

file_list = config['paths']['file_list']
af = config['paths']['final_corpus']
dir = config['paths']['output']+'/classics_paper'
filterStopWords=config['params']['filter_stop_words']

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]

for file in os.listdir(dir):
	os.remove('%s/%s'%(dir,file))

for idx,record in enumerate(files):
	file = '%s/%s'%(af,record[h['Tokenized file']].value)
	aesthetics = record[h['Aesthetics']].value
	foolP = str(record[h['Fool: Period']].value)
	foolA = str(record[h['Fool: Author']].value)
	wy = str(record[h['Date']].value)
	sys.stdout.write("\r\033[KProcessing %s"%record[h['Tokenized file']].value)
	sys.stdout.flush()
	finalTxt = open(file, 'r').read().replace("\n", "").replace(" ", "")
	finalTxt = re.sub('.*?<body>(.*?)</body>.*', r'\1', finalTxt)
	finalTxt = re.sub('<punct.*?/>', '', finalTxt)
	finalTxt = re.sub('<sentenceid=".*?"location=".*?"/?>', '['+wy+']\t', finalTxt)
	finalTxt = re.sub('<word.*?lemmaid="(.*?)".*?</word>', r'*\1*', finalTxt)
	finalTxt = re.sub('</sentence>', '\n', finalTxt)
	if filterStopWords == "True":
		for stop in STOPS_LIST_ID:
			finalTxt = re.sub('\*%s\*'%stop, '', finalTxt)
		finalTxt = re.sub('.*?\t\n','',finalTxt)
	finalTxt = finalTxt.replace('**', ' ').replace('*', '')
	finalTxt = re.sub('^-?\d+\n','',finalTxt)
	conv_text = []
	conv_text_l = []
	for x in finalTxt.split():	
		if re.search('\[.*?\]', x) != None:
			conv_text.append("\n"+x)
			conv_text_l.append("\n"+x)
		else:
			conv_text.append(x)
			conv_text_l.append(greekLemmata[x]['lemma'])
	finalTxt = ' '.join(conv_text_l).replace('] ', ']\t')
	finalTxt = finalTxt.replace('[', '').replace(']', '').strip()
	finalTxt_ids = ' '.join(conv_text).replace('] ', ']\t')
	finalTxt_ids = finalTxt_ids.replace('[', '').replace(']', '').strip()
	if aesthetics == 'x':
		finalTxt_fP = re.sub('-?\d+(?=\t)',foolP,finalTxt).strip()
		finalTxt_fA = re.sub('-?\d+(?=\t)',foolA,finalTxt).strip()
		finalTxt_fP_ids = re.sub('-?\d+(?=\t)',foolP,finalTxt_ids).strip()
		finalTxt_fA_ids = re.sub('-?\d+(?=\t)',foolA,finalTxt_ids).strip()
		fullText=open('%s/aesth_periods_forms.txt'%dir, 'a+')
		fullText.write(finalTxt_fP+'\n')
		fullText.close()
		fullText=open('%s/aesth_periods_ids.txt'%dir, 'a+')
		fullText.write(finalTxt_fP_ids+'\n')
		fullText.close()
		fullText=open('%s/aesth_authors_forms.txt'%dir, 'a+')
		fullText.write(finalTxt_fA+'\n')
		fullText.close()
		fullText=open('%s/aesth_authors_ids.txt'%dir, 'a+')
		fullText.write(finalTxt_fA_ids+'\n')
		fullText.close()
		del finalTxt_fP, finalTxt_fP_ids, finalTxt_fA, finalTxt_fA_ids
	else:
		fullText=open('%s/non_aesth_forms.txt'%dir, 'a+')
		fullText.write(finalTxt+'\n')
		fullText.close()
		fullText=open('%s/non_aesth_ids.txt'%dir, 'a+')
		fullText.write(finalTxt_ids+'\n')
		fullText.close()
		del finalTxt, finalTxt_ids
		
print('\n###########\n#All done!#\n###########\n')