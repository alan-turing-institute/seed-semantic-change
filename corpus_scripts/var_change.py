import os
import re
import configparser
import numpy as np
import pandas as pd
import sys
from openpyxl import load_workbook
from scipy import stats

verbose = False

if 'verbose' in sys.argv: verbose = True

os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')
dir = config['paths']['output']
resources = config['paths']['file_list']

wb = load_workbook('%s/word_senses.xlsx'%resources)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h_sense = {cell.value : n for n, cell in enumerate(headers[0])}
data = ws['A2:E23']

word_senses = {}
for record in data:
	word = record[h_sense['TERM']].value
	sense_id = record[h_sense['SENSE LSJ']].value
	sense = record[h_sense['SENSE']].value
	word_senses.setdefault(sense_id, sense)

if os.path.isfile('%s/change_data_macro.txt'%dir): os.remove('%s/change_data_macro.txt'%dir)
change_file = open('%s/change_data_macro.txt'%dir, 'a')
if os.path.isfile('%s/variation_data_macro.txt'%dir): os.remove('%s/variation_data_macro_macro.txt'%dir)
variation_file = open('%s/variation_data_macro.txt'%dir, 'a')

input_files = ['15281', '69419']

scores={}
for file in input_files:
	lf = open('%s/senses_%s.txt'%(dir,file), 'r')
	for line in lf:
		if re.search('[\-\d]', line[0]) == None:
			continue
		sense = line.split('\t')[-3]
		if sense =='w':
			continue
		genre = line.split('\t')[1].strip()
		year = int(line.split('\t')[0])
		if year >= -800 and year <= -600:
			century = '-7'
		elif year > -600 and year <= -500:
			century = '-6'
		elif year > -500 and year <= -400:
			century = '-5'
		elif year > -400 and year <= -300:
			century = '-4'
		elif year > -300 and year <= -200:
			century = '-3'
		elif year > -200 and year <= -100:
			century = '-2'
		elif year > -100 and year <= -1:
			century = '-1'
		elif year >= 0 and year < 100:
			century = '1'
		elif year >= 100 and year < 200:
			century = '2'
		elif year >= 200 and year < 300:
			century = '3'
		elif year >= 300 and year < 400:
			century = '4'
		elif year >= 400 and year < 500:
			century = '5'
		scores.setdefault((sense,genre,century),0)
		scores[(sense,genre,century)]+=1

centuries=['-7','-6','-5','-4','-3','-2','-1','1','2','3','4','5']
genres = set()
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find('-')]
	genres.add(genre)
	
# { sense0 {
	#scores for t0...tn: [] | scores for t0...tn for genre0: [] | scores for t0...tn for genre...n: []
# } }

word_scores = {}
for (sense,genre,century),score in scores.items():
	word_scores.setdefault(sense,{}).setdefault(century,{}).setdefault(genre,0)
	word_scores[sense][century][genre]+=score

for century in centuries:
	for genre in genres:
		for sense,data in word_scores.items():
			word_scores[sense].setdefault(century,{}).setdefault(genre,0)

for sense,data in word_scores.items():
	table = {} #col = [ ]
	for century in centuries:
		line = []
		for genre,score in data[century].items():
			line.append(score)
			table.setdefault(genre,[]).append(score)
		table.setdefault('TOTAL', []).append(sum(line))
	table=pd.DataFrame(data=table, index=centuries)
	fields = [x for x in table if x != 'TOTAL']
	fields.append('TOTAL')
	table=table[fields]
	table_tots=table[[x for x in table]].sum()
	table=table.transpose()
	table['TOTAL']=table_tots
	print('\033[1m'+100*'#'+'\n'+'#   '+sense+(95-len(sense))*' '+'#\n'+100*'#'+'\033[0m\n')
	variation_file.write(100*'#'+'\n#   '+sense+(95-len(sense))*' '+'#\n'+100*'#'+'\n')
	with pd.option_context('expand_frame_repr', False):
		if verbose ==True: print(table,'\n')
		variation_file.write('\n'+re.sub(' {2,}','\t',table.to_string())+'\n')
	
	#subplots
	corr_ranking = []
	for x in genres:
		if verbose ==True: print(sense, '~', x)
		variation_file.write('\n'+sense+' ~ '+x)
		if verbose ==True: print(table.loc[[x,'TOTAL'],[x for x in centuries]])
		variation_file.write('\n'+table.loc[[x,'TOTAL'],[x for x in centuries]].to_string())
		new_matrix = table.loc[[x,'TOTAL'],[x for x in centuries]].values.tolist()
		if sum(new_matrix[0]) == 0:
			r=(np.nan,np.nan)
		else: 
			r = stats.spearmanr(new_matrix[0], new_matrix[1])
		if verbose ==True: print(r, '\n')
		variation_file.write('\n'+str(r)+'\n')
		corr_ranking.append(('%s ~ %s'%(sense,x),r[0],r[1]))
	ranking=pd.DataFrame([x[1:] for x in corr_ranking], columns=['Spearman\'s ρ','p'],index=[x[0] for x in corr_ranking])
	ranking=ranking.reindex(ranking[['Spearman\'s ρ']].abs().sort_values(by=['Spearman\'s ρ'], ascending=False).index)
	print(ranking,'\n')
	variation_file.write('\n'+ranking.to_string()+'\n\n')
variation_file.close()
		
#change points
#conditions
# (1) ∀t(x) if s(a) of w(b) is first attested in G(α), ..., G(ω) 
# (2) & G(α) or ... or G(ω) is attested at t(y) : y < x  
# (3) & w(b) is attested in G(α) or ... or G(ω) at t(y) : y < x
# (2) implicata da (3): 2 = F => 3 = F; 2 = T ¬=> 3 = T; 3 = T => 2 = T; 3 = F ¬=> 2 = F; siccome (2) & (3) devono essere T, basta che sia soddisfatta 3

#Cosa voglio:
#tabella = parola
#		|prima attestazione	|	generi
#————————————————————————————————————————————————————————
#senso1	|	secolo			|	[genere(a) ... genere(n)]

change_scores = {}
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find('-')]
	change_scores.setdefault(word,{}).setdefault(sense,{}).setdefault(int(century),set()).add(genre)
	
words_in_genres = {}
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find('-')]
	words_in_genres.setdefault(word,{}).setdefault(int(century),set()).add(genre)
	
	#each word must be a new table (so, new dictionary entry)
	#each sense must be a new line, containing only data on the first attestation(s): century, genres
	#so: sense \ century \ genres;
	#then for each sense, sort by century and only take the earliest one
	#century \ genres can be a tuple (cent, [list of genres])
	#senses should be a dictionary: {word1 : {sense1 : [{cent1: {set of genres}}]}}
	#then turn this into a dataframe
pd.set_option('display.max_colwidth', -1)
#now turn this into a dataframe per word
for word, data in change_scores.items():
	print('\033[1m'+40*'#'+'\n'+'#   '+word+(35-len(word))*' '+'#\n'+40*'#'+'\033[0m\n')
	change_file.write(40*'#'+'\n#   '+word+(35-len(word))*' '+'#\n'+40*'#'+'\n')
	for sense, centuries in data.items():
		century_list = []
		for century, genres in centuries.items():
			new_tuple = (century, [x for x in genres])
			century_list.append(new_tuple)
		century_list=sorted(century_list, key=lambda cent: cent[0])
		first_cent = century_list[0][0]
		first_genres = century_list[0][1]
		cent_att_before = [cent for cent in words_in_genres[word].keys() if cent < first_cent]
		genr_att_before = set()
		for cent in cent_att_before:
			for w in (words_in_genres[word][cent]):
				genr_att_before.add(w)
		word_attested_before=len(cent_att_before)>0
		change = False
		if len([x for x in first_genres if x in genr_att_before]) > 0: change=True
		data[sense]=(first_cent,first_genres,word_attested_before,genr_att_before,change)
	table=pd.DataFrame(data=data, index=['First attestation', 'First genres', 'Word attested before','Genres attested before','Innovation']).transpose()
	with pd.option_context('expand_frame_repr', False):
		change_file.write('\n'+re.sub(' {2,}','\t',table.to_string())+'\n\n')
		print(table,'\n')
change_file.close()
