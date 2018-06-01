import re
import os
import sys
import lxml.etree as document
from utf2beta import convertUTF
from beta2utf import convertBeta
from grkFrm import greekForms 
from grkLemmata import greekLemmata
from openpyxl import load_workbook
import configparser
import json

os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')
file_list = config['paths']['file_list']
tf = config['paths']['tokenized']
output = config['paths']['output']
parser = document.XMLParser(remove_blank_text=True)

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]

result_file = open('%s/ambiguity_stats.txt'%output, 'w+')

def dbl_output(*args, **kwargs):
	global result_file
	if len(kwargs.items()) == 0:
		sep = ' '
	else:
		sep = kwargs['sep']
	if len(args) == 0:
		print()
		result_file.write('\n')
	else:
		print(sep.join([str(x) for x in args[0:]]))
		result_file.write('%s\n'%sep.join([str(x) for x in args[0:]]))

os.system("clear && printf '\e[3J'")

#######################################
#compute ambiguous forms in dictionary#
#######################################

mlen = 0
amb_types_multi_pos = 0
posDictOneLemma = 0
for key,items in greekForms.items():
	#a=set()
	#for lemma in items:
#		a.add(json.dumps(lemma, sort_keys=True))
	#if len(a)!=len(items): print('\n',key,'\t',items)
	pos_one_lemma_dict = 0
	if len(items)>1:
		mlen+=1
		pos_in_type={}
		for lemma in items:
			pos_in_type[greekLemmata[lemma['l']]['pos']]=pos_in_type.setdefault(greekLemmata.get(lemma['l'],'unknown').get('pos','unknown'),0)+1
		if len(pos_in_type.items()) > 1:
			amb_types_multi_pos += 1
			for pos,count in pos_in_type.items():
				if count == 1:
					pos_one_lemma_dict = 1
	posDictOneLemma += pos_one_lemma_dict
	
####################################
#compute ambiguous tokens in corpus#
####################################

def cleanWords(word):
	word = re.sub('[^a-z\(\)\\\/\*\|=\+\'0-9\"]*', '', word)

	#turn graves into acutes, lowercase betacode, remove clitic accent, escape diacritics
	wordForm = word.replace("2", "").replace("3", "")
	word = word.replace('\\', '/').lower()
	word = re.sub(r"\*([aehiouw])([\(\)/=]+)",r"*\2\1", word) #parole registrate maiuscole: *DIACRlettere; parole minuscole: *parola normale, quindi * + e)/ etc.
	word = re.sub(r"([\(\)/=]+)(\*)([aehiouw])",r"\2\1\3", word)
	word = re.sub(r"\|([/\(\)])",r"\g<1>|", word)
	howmanyAccents = len(re.findall("[/=]", word))
	if howmanyAccents > 1:
		word=word[::-1].replace("/", "", 1)[::-1]
	return word
	
def lookup(word):
	word = cleanWords(word)
	try:
		entry=greekForms[word]
	#no match			
	except KeyError:
		word = word.replace("+", "")
		word = re.sub(r"\*([\(\)/=]+)([aehiouw])",r"*\2\1", word)
		word2 = word.replace("'", "")
		notFound = False
		try:
			entry=greekForms[word]
		except KeyError:
			notFound = True
		else:
			return {'is_unknown':False, 'entry':entry}
		if notFound == True:
			word = word.replace("*", "")		
			try:
				entry=greekForms[word]
			except KeyError:
				word = re.sub('^c(?=u[nmg])','s', word)
				try:
					entry=greekForms[word]
				except KeyError:
					word = word.replace("'", "")
					try:
						entry=greekForms[word]
					except KeyError:
						word = re.sub("\)$", "'",word)
						try:
							entry=greekForms[word]
						except KeyError:
							word = re.sub("'$", "",word)
							try:
								entry=greekForms[word]
							except KeyError:
								word = re.sub("^\(", "",word)
								try:
									entry=greekForms[word]
								except KeyError:
									word = re.sub("\)$", "",word)
									try:
										entry=greekForms[word]
									except KeyError:
										word = re.sub("'", "",word)
										try:
											entry=greekForms[word]
										except KeyError:
											return {'is_unknown':True}
										else:
											return {'is_unknown':False, 'entry':entry}
									else:
										return {'is_unknown':False, 'entry':entry}
								else:
									return {'is_unknown':False, 'entry':entry}
							else:
								return {'is_unknown':False, 'entry':entry}
						else:
							return {'is_unknown':False, 'entry':entry}
					else:
						return {'is_unknown':False, 'entry':entry}
				else:
					return {'is_unknown':False, 'entry':entry}
			else:
					return {'is_unknown':False, 'entry':entry}
	else:
		return {'is_unknown':False, 'entry':entry}

def process(lemmata, word):
	#each wordform corresponds to a list of lemmata
	global wml
	global wml_mPos
	global wTypesml
	global wTypesTot
	global wTypesmlmpos
	global posTokensOneLemma
	global posTypesOneLemma

	wTypesTot.add(word.get('form'))
	if len(lemmata)>1:
		wml+=1
		wTypesml.add(word.get('form'))
		pos_in_word={}
		for lemma in lemmata:
			curr_pos = greekLemmata[lemma['l']]['pos']
			pos_in_word[curr_pos]=pos_in_word.setdefault(curr_pos, 0)+1
		pos_one_lemma = 0
		pos_one_lemma_tokens = 0
		if len(pos_in_word.items()) > 1:
			wml_mPos += 1
			wTypesmlmpos.add(word.get('form'))
			for pos,count in pos_in_word.items():
				if count == 1:
					pos_one_lemma+=1
					pos_one_lemma_tokens=1
					posTypesOneLemma.add(word.get('form'))
		posTokensOneLemma += pos_one_lemma_tokens
	

wml = 0 		#words with multiple lemmata
wml_mPos = 0	#word with multiple lemmata disambiguable by part of speech (cases: 1*POS1,1*POS2,1*...,1*POSn = T; >1*POS1,...,1*POSn = T only if tagger outputs POS occurring only once; >1*POS1,...,>1*POSn = F
wTypesml=set()
wTypesTot=set()
wTypesmlmpos=set()
totTokens = 0
posTokensOneLemma = 0
posTypesOneLemma = set()
for record in files:
	file = '%s/%s'%(tf,record[h['Tokenized file']].value)
	progPercent=0
	fileName = record[h['Tokenized file']].value
	parse = document.parse(file, parser)
	words = parse.xpath('//word')
	total = len(words)
	totTokens += total
	for idx, word in enumerate(words):
		progPercent = int(idx*30/total)*'='+(29-int(idx*30/total))*' '
		sys.stdout.write("\r\033[K%s - %s, progress: [%s] %s%%"%(record[h['Author']].value,record[h['Work']].value[:18]+'...' if len(record[h['Work']].value)>15 else record[h['Work']].value,progPercent,str(int(idx*100/total))))
		sys.stdout.flush()
		#get word form
		form = word.get("form")
		isunknown = False
		#lookup in dictionary
		if lookup(form)['is_unknown'] == True:
		#####test lacunae (merge with following words)#####
			try:
				next_word = words[idx+1]
			except:
				isunknown = True
			else:
				next_form = next_word.get("form")
				iter = 0
				while True:
					iter += 1
					previous_word = words[idx-iter]
					if previous_word.get('toremove') == None:
						break
				previous_form = previous_word.get("form")
				merged_form = ''
				if word.get('lacuna') != None:
					merged_form = '%s%s'%(form, next_form)
					if lookup(merged_form)['is_unknown'] == True:
						try:
							next_next_word = words[idx+2]
						except:
							isunknown = True
						else:
							next_next_form = next_next_word.get("form")
							if next_next_word.get('lacuna') != None:
								merged_form='%s%s'%(merged_form, next_next_form)
								if lookup(merged_form)['is_unknown'] == True:
									try:
										next_next_next_word = words[idx+3]
									except:
										isunknown = True
									else:
										next_next_next_form = next_next_next_word.get("form")
										merged_form = '%s%s'%(merged_form, next_next_next_form)
										if lookup(merged_form)['is_unknown'] == True:
											isunknown = True
										else:
											word.set('form', merged_form)
											process(lookup(merged_form)['entry'], word)
											next_word.set('toremove', 'true')
											next_next_word.set('toremove', 'true')
											next_next_next_word.set('toremove', 'true')
								else:
									word.set('form', merged_form)
									process(lookup(merged_form)['entry'], word)
									next_word.set('toremove', 'true')
									next_next_word.set('toremove', 'true')
							else:
								isunknown = True
					else:
						word.set('form', merged_form)
						process(lookup(merged_form)['entry'], word)
						next_word.set('toremove', 'true')
				else:
					if next_word.get('lacuna') != None:
						merged_form = '%s%s'%(form, next_form)
						if lookup(merged_form)['is_unknown'] == True:
							try:
								next_next_word = words[idx+2]
							except:
								isunknown = True
							else:
								next_next_form = next_next_word.get("form")
								merged_form='%s%s'%(merged_form, next_next_form)
								if lookup(merged_form)['is_unknown'] == True:
									try:
										next_next_next_word = words[idx+3]
									except:
										isunknown = True
									else:
										next_next_next_form = next_next_next_word.get("form")
										merged_form = '%s%s'%(merged_form, next_next_next_form)
										if lookup(merged_form)['is_unknown'] == True:
											isunknown = True
										else:
											word.set('form', merged_form)
											process(lookup(merged_form)['entry'], word)
											next_word.set('toremove', 'true')
											next_next_word.set('toremove', 'true')
											next_next_next_word.set('toremove', 'true')
								else:
									word.set('form', merged_form)
									word.set('lacuna', 'true')
									process(lookup(merged_form)['entry'], word)
									next_word.set('toremove', 'true')
									next_next_word.set('toremove', 'true')
						else:
							word.set('form', merged_form)
							word.set('lacuna', 'true')
							process(lookup(merged_form)['entry'], word)
							next_word.set('toremove', 'true')
					else:
						if form[-1] == '-':
							merged_form = '%s%s'%(form, next_form)
							if lookup(merged_form)['is_unknown'] == True:
								isunknown = True
							else:
								word.set('form', merged_form)
								process(lookup(merged_form)['entry'], word)
								next_word.set('toremove', 'true')
						else:
							isunknown = True
		#####end test lacunae (1)#####
		#####test lacunae (merge with previous word and next word)#####
			if isunknown == True and (word.get('lacuna') != None or previous_word.get('lacuna') != None):
				merged_form='%s%s'%(previous_form, form)
				if lookup(merged_form)['is_unknown'] == True:
					merged_form='%s%s'%(merged_form, next_form)
					if lookup(merged_form)['is_unknown'] == True:
						isunknown = True
					else:
						previous_word.set('form', merged_form)
						previous_word.set('lacuna', 'true')
						process(lookup(merged_form)['entry'], previous_word)
						word.set('toremove', 'true')
						next_word.set('toremove', 'true')
				else:
					previous_word.set('form', merged_form)
					previous_word.set('lacuna', 'true')
					process(lookup(merged_form)['entry'], previous_word)
					word.set('toremove', 'true')
		#####end test lacunae (2)#####

			if isunknown == True:
				document.SubElement(word, "lemma").set("id", "unknown")
		else:
			process(lookup(form)['entry'], word)

	#remove joined words
	toremove=parse.xpath('//word[@toremove]')
	for nTr in toremove:
		nTr.getparent().remove(nTr)

#########		
#results#		
#########
print()
dbl_output('α.a.1',mlen,'out of',len(greekForms),'forms in dictionary can belong to more than one lemma')
dbl_output('α.a.2',len(wTypesml),'types with multiple lemmata in corpus out of', len(wTypesTot))
dbl_output('α.b',wml, 'tokens with multiple lemmata in corpus out of', totTokens)
dbl_output('β.a.1',amb_types_multi_pos,'out of',len(greekForms),'forms in dictionary can belong to more than one lemma and such lemmas belong to more than 1 PoS')
dbl_output('β.a.2',len(wTypesmlmpos),'out of',len(wTypesTot),'types in corpus can belong to more than one lemma and such lemmas belong to more than 1 PoS')
dbl_output('β.b',wml_mPos,'out of',totTokens,'tokens in corpus can belong to more than one lemma and such lemmas belong to more than 1 PoS')
dbl_output('γ.a.2',posDictOneLemma,'out of',len(greekForms),'forms in dictionary can belong to more than one lemma and such lemmas belong to more than 1 PoS, and have at least one PoS represented by only one lemma')
dbl_output('γ.a.2',len(posTypesOneLemma),'out of',len(wTypesTot),'types in corpus can belong to more than one lemma and such lemmas belong to more than 1 PoS, and have at least one PoS represented by only one lemma')
dbl_output('γ.b',posTokensOneLemma,'out of',totTokens,'tokens in corpus can belong to more than one lemma and such lemmas belong to more than 1 PoS, and have at least one PoS represented by only one lemma')

#β.a. how many ambiguous types belong to >1 PoS
#		β.a.1	in dictionary
#		β.a.2	in corpus	wTypesmlmpos
#	β.b. how many ambiguous tokens belong to >1 PoS in corpus wml_mPos
#	γ.a. how many PoS in types (β.a) are represented by only one lemma
#		γ.a.1	in dictionary
#		γ.a.2	in corpus
#	γ.b.	how many PoS in tokens (β.b) are represented by only one lemma in corpus



#	record[h['Disambiguable words']].value = wml_mPos