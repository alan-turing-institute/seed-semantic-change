import os
import configparser
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.realpath(__file__)))
os.system("clear && printf '\e[3J'")
config = configparser.ConfigParser()
config.read('config.ini')
treetagger = config['paths']['treetagger']

#Tagging test set

print('Tagging test set')
os.system("{tt_folder}/tree-tagger {curr_folder}/TreeTaggerData/ancient_greek.dat {curr_folder}/TreeTaggerData/test_set.txt {curr_folder}/TreeTaggerData/test_set_tagged.txt -token -lemma".format(tt_folder=treetagger, curr_folder=os.path.dirname(os.path.realpath(__file__))))

a_lines=[x for x in open('TreeTaggerData/test_set_tagged.txt', 'r')]
c_lines=[x for x in open('TreeTaggerData/test_set_benchmark.txt', 'r')]
open('TreeTaggerData/test_set_results.txt', 'w')
output = open('TreeTaggerData/test_set_results.txt', 'a')
trues = 0
POS_list = {} #number of forms tagged as N by TreeBanker
POS_list_benchmark = {} #number of forms annotated as N in the treebank
POS_list_correct = {} #number of forms correctly tagged as N
for idx, x in enumerate(range(0,len(a_lines))):
	#output.write('%s\t%s\t%s\n'%(a_lines[idx].strip(),c_lines[idx].strip(),a_lines[idx] == c_lines[idx]))
	POS_a=a_lines[idx].strip().split('\t')[1]
	POS_c=c_lines[idx].strip().split('\t')[1]
	POS_list.setdefault(POS_a,0)
	POS_list[POS_a]+=1
	POS_list_benchmark.setdefault(POS_c,0)
	POS_list_benchmark[POS_c]+=1
	POS_list_correct.setdefault(POS_a,0)
	if a_lines[idx].strip().split('\t')[0:2] == c_lines[idx].strip().split('\t')[0:2]:
		trues+=1
		POS_list_correct[POS_a]+=1
percent_true = round(trues*100/len(a_lines),2)
output.write('======================== GENERAL RESULTS ========================')
output.write('\n\nAccuracy: %d%%'%percent_true)


#Precision = (number of forms correctly tagged as N)/(number of forms tagged as N)
#Recall = (number of forms correctly tagged as N)/(number of Ns in the gold standard)
#F-score = 2*(precision*recall)/(precision +recall)

for POS,count in POS_list.items():
	tagged_as_N=count
	correctly_tagged=POS_list_correct.get(POS,0)
	count_in_gold_standard=POS_list_benchmark.get(POS,0)
	try:
		precision = correctly_tagged/tagged_as_N
	except:
		precision = 'N/A'
	try:
		recall = correctly_tagged/count_in_gold_standard
	except:
		recall = 'N/A'
	try:
		Fscore = 2*(precision*recall)/(precision+recall)
	except:
		Fscore = 'N/A'
	output_string = '\n\n'
	output_string += 30*'#'
	output_string +=('\n# %s%s#'%(POS,(27-len(POS))*' '))
	output_string +=('\n'+30*'#')
	output_string +=('\n\tnumber of forms tagged as %s: %d'%(POS, tagged_as_N))
	output_string +=('\n\tnumber of forms annotated as %s in treebank: %d'%(POS, count_in_gold_standard))
	output_string +=('\n\tnumber of forms correctly tagged: %d'%correctly_tagged)
	output_string +=('\n\tPrecision:\t%s'%precision)
	output_string +=('\n\tRecall:\t\t%s'%recall)
	output_string +=('\n\tF-score:\t%s'%Fscore)
	output.write(output_string)
	
#############	
	
#Tag test set files and retrieve metadata
print("Tagging files")

file_list = config['paths']['file_list']
wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]

authorResults = {}
genreResults = {}
POS_list_authors = {}
POS_list_genres = {}
POS_list_benchmark_authors = {}
POS_list_benchmark_genres = {}
POS_list_correct_authors = {}
POS_list_correct_genres = {}

for filename in os.listdir('TreeTaggerData/TestSetFiles'):
	if filename.find('benchmark')==-1 and filename.find('tagged')==-1 and filename[-3:]=='txt':
		author = filename[3:7]
		work = filename[11:14]
		for record in files:
			if record[h['TLG Author']].value == author and record[h['TLG ID']].value == work: #retrieve metadata
				author = record[h['Author']].value
				work = record[h['Work']].value
				genre = record[h['Genre']].value
				break			
		#tag texts
		os.system("{tt_folder}/tree-tagger {curr_folder}/TreeTaggerData/ancient_greek.dat {curr_folder}/TreeTaggerData/TestSetFiles/{filename} {curr_folder}/TreeTaggerData/TestSetFiles/{filename}_tagged.txt -token -lemma".format(tt_folder=treetagger, curr_folder=os.path.dirname(os.path.realpath(__file__)), filename=filename))
		
		#compare tagged with benchmark
		a_lines=[x for x in open('TreeTaggerData/TestSetFiles/%s_tagged.txt'%filename, 'r')]
		c_lines=[x for x in open('TreeTaggerData/TestSetFiles/%s.txt'%filename.replace('.txt','_benchmark'), 'r')]
		
		trues = 0
		for idx, x in enumerate(range(0,len(a_lines))):
			if a_lines[idx].strip().split('\t')[0:2] == c_lines[idx].strip().split('\t')[0:2]:
				trues+=1
		authorResults.setdefault(author, [0,0])
		authorResults[author][0]+=len(a_lines)
		authorResults[author][1]+=trues #forms correctly tagged
		genreResults.setdefault(genre, [0,0])
		genreResults[genre][0]+=len(a_lines)
		genreResults[genre][1]+=trues #forms correctly tagged
		
		POS_list_authors.setdefault(author,{})
		POS_list_genres.setdefault(genre,{})
		POS_list_benchmark_authors.setdefault(author,{})
		POS_list_benchmark_genres.setdefault(genre,{})
		POS_list_correct_authors.setdefault(author,{})
		POS_list_correct_genres.setdefault(genre,{})
		
		for idx, x in enumerate(range(0,len(a_lines))):

			POS_a=a_lines[idx].strip().split('\t')[1]
			POS_c=c_lines[idx].strip().split('\t')[1]
			POS_list_authors[author].setdefault(POS_a,0)
			POS_list_genres[genre].setdefault(POS_a,0)
			POS_list_authors[author][POS_a]+=1
			POS_list_genres[genre][POS_a]+=1
			POS_list_benchmark_authors[author].setdefault(POS_c,0)
			POS_list_benchmark_genres[genre].setdefault(POS_c,0)
			POS_list_benchmark_authors[author][POS_c]+=1
			POS_list_benchmark_genres[genre][POS_c]+=1
			POS_list_correct_authors[author].setdefault(POS_a,0)
			POS_list_correct_genres[genre].setdefault(POS_a,0)
			if a_lines[idx].strip().split('\t')[0:2] == c_lines[idx].strip().split('\t')[0:2]:
				trues+=1
				POS_list_correct_authors[author][POS_a]+=1
				POS_list_correct_genres[genre][POS_a]+=1
		

output.write('\n\n======================= RESULTS BY AUTHOR =======================')
for author,values in authorResults.items():
	output_string = '\n\n'
	output_string += 30*'#'
	output_string +=('\n# %s%s#'%(author,(27-len(author))*' '))
	output_string +=('\n'+30*'#')
	output_string +=('\n\n\tAccuracy: %d%%'%round(values[1]*100/values[0], 2))
	output.write(output_string)
	for POS,count in POS_list_authors[author].items():
		tagged_as_N=count
		correctly_tagged=POS_list_correct_authors[author].get(POS,0)
		count_in_gold_standard=POS_list_benchmark_authors[author].get(POS,0)
		try:
			precision = correctly_tagged/tagged_as_N
		except:
			precision = 'N/A'
		try:
			recall = correctly_tagged/count_in_gold_standard
		except:
			recall = 'N/A'
		try:
			Fscore = 2*(precision*recall)/(precision+recall)
		except:
			Fscore = 'N/A'
		output_string =('\n\n\t%s in %s'%(POS,author))
		output_string +=('\n\t\tnumber of forms tagged as %s: %d'%(POS, tagged_as_N))
		output_string +=('\n\t\tnumber of forms annotated as %s in treebank: %d'%(POS, count_in_gold_standard))
		output_string +=('\n\t\tnumber of forms correctly tagged: %d'%correctly_tagged)
		output_string +=('\n\t\tPrecision:\t%s'%precision)
		output_string +=('\n\t\tRecall:\t\t%s'%recall)
		output_string +=('\n\t\tF-score:\t%s'%Fscore)
		output.write(output_string)
				
output.write('\n\n=========================== RESULTS BY GENRE ===========================')
for author,values in genreResults.items():
	output_string += '\n\n'
	output_string += 30*'#'
	output_string +=('\n# %s%s#'%(author,(27-len(author))*' '))
	output_string +=('\n'+30*'#')
	output_string +=('\n\n\tAccuracy: %d%%'%round(values[1]*100/values[0], 2))
	output.write(output_string)
	for POS,count in POS_list_genres[genre].items():
		tagged_as_N=count
		correctly_tagged=POS_list_correct_genres[genre].get(POS,0)
		count_in_gold_standard=POS_list_benchmark_genres[genre].get(POS,0)
		try:
			precision = correctly_tagged/tagged_as_N
		except:
			precision = 'N/A'
		try:
			recall = correctly_tagged/count_in_gold_standard
		except:
			recall = 'N/A'
		try:
			Fscore = 2*(precision*recall)/(precision+recall)
		except:
			Fscore = 'N/A'
		output_string =('\n\n\t%s in %s'%(POS,author))
		output_string +=('\n\t\tnumber of forms tagged as %s: %d'%(POS, tagged_as_N))
		output_string +=('\n\t\tnumber of forms annotated as %s in treebank: %d'%(POS, count_in_gold_standard))
		output_string +=('\n\t\tnumber of forms correctly tagged: %d'%correctly_tagged)
		output_string +=('\n\t\tPrecision:\t%s'%precision)
		output_string +=('\n\t\tRecall:\t\t%s'%recall)
		output_string +=('\n\t\tF-score:\t%s'%Fscore)
		output.write(output_string)
		
output.close()
os.system("clear && printf '\e[3J'")
print(open('TreeTaggerData/test_set_results.txt','r').read())