import os
import re
import configparser
from openpyxl import load_workbook
from grkLemmata import greekLemmata
import operator

os.chdir(os.path.dirname(os.path.realpath(__file__)))
os.system("clear && printf '\e[3J'")
config = configparser.ConfigParser()
config.read('config.ini')
source_path = config['paths']['output']
resources = config['paths']['file_list']
window = int(config['params']['window'])

wb = load_workbook('%s/word_senses.xlsx'%resources)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
data = ws['A2:D22']

print('Bear with me a sec...')

target_words = {}
for record in data:
	target_words.setdefault(record[h['TERM ID']].value, {})

full_corpus = open('%s/%s'%(source_path, 'fctt.txt'), 'r')
sentences = [x for x in full_corpus]
for sentence in sentences:
	try:
		words = sentence.split('\t')[1].split()
		for target_word,context_words in target_words.items():
			if target_word in words:
				indexTW = words.index(target_word)
				start = 0
				finish = len(words)
				if len(words)>2*window+1:
					start = indexTW-window
					finish = indexTW+window+1
				for word in words[start:finish]:
					context_words.setdefault(word.strip(),0)
					context_words[word.strip()]+=1
	except:
		continue
	
for word,list in target_words.items():
	open('%s/collocates_%s_[%s].csv'%(source_path,greekLemmata[word]['lemma'],word), 'w')
	output=open('%s/collocates_%s_[%s].csv'%(source_path,greekLemmata[word]['lemma'],word), 'a')
	#retrieve senses from file
	senses = []
	for record in data:
		if record[h['TERM ID']].value == word:
			senses.append(record[h['SENSE']].value)
	output.write('Target word,Target word ID,Collocate,Collocate ID,Collocate frequency,%s,generic\n'%','.join(senses))
	print(greekLemmata[word]['lemma'])
	sorted_list = sorted(list.items(), key=operator.itemgetter(1), reverse=True)
	for collocate, frequency in sorted_list:
		if collocate != word:
			print('\t%s: %s'%(greekLemmata[collocate]['lemma'], frequency))
			output.write('%s,%s,%s,%s,%s,\n'%(greekLemmata[word]['lemma'],word,greekLemmata[collocate]['lemma'],collocate,frequency))

print('All done!')
