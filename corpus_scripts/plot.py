import os
import re
import configparser
import matplotlib.pyplot as plot
import matplotlib.colors as colors
import numpy
import sys
from openpyxl import load_workbook
from statsmodels.stats import proportion

import matplotlib as plt
plt.rcParams.update({'figure.max_open_warning': 0})

os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')
dir = config['paths']['output']
resources = config['paths']['file_list']

wb = load_workbook('%s/word_senses.xlsx'%resources)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h_sense = {cell.value : n for n, cell in enumerate(headers[0])}
data = ws['A2:F23']

sep = ':'

word_senses = {}
for record in data:
	word = record[h_sense['TERM']].value
	sense_id = record[h_sense['MERGED SENSE']].value
	sense = record[h_sense['MERGED SENSE']].value.split(':')[1]
	word_senses.setdefault(sense_id, sense)

input_files = ['15281', '69419']
scores={}
colour={}
colours={hex for name, hex in colors.cnames.items()}
#removing colours too bright/dark to display
toPop=[]
for color in colours:
	lum = 0.2126*int(color[1:3],16)+0.7152*int(color[3:5],16)+0.0722*int(color[5:7],16)
	if lum > 175 or lum < 120:
		toPop.append(color)
for color in toPop:
	colours.remove(color)

for file in input_files:
	lf = open('%s/senses_%s.txt'%(dir,file), 'r')
	for line in lf:
		if re.search('[\-\d]', line[0]) == None:
			continue
		sense = line.split('\t')[-4]
		if sense =='wrong':
			continue
		genre = line.split('\t')[1]
		year = int(line.split('\t')[0])
		if year >= -800 and year <= -700:
			century = '-8'
		elif year > -700 and year <= -600:
			century = '-7'
		elif year > -600 and year <= -500:
			century = '-6'
		elif year > -500 and year <= -400:
			century = '-5'
		elif year > -400 and year <= -300:
			century = '-4'
		elif year > -300 and year <= -200:
			century = '-3'
		elif year > -200 and year <= -100:
			century = '-2'
		elif year > -100 and year <= -1:
			century = '-1'
		elif year >= 0 and year < 100:
			century = '1'
		elif year >= 100 and year < 200:
			century = '2'
		elif year >= 200 and year < 300:
			century = '3'
		elif year >= 300 and year < 400:
			century = '4'
		elif year >= 400 and year < 500:
			century = '5'
		scores.setdefault((sense,genre,century),0)
		scores[(sense,genre,century)]+=1
		try:
			colour[sense]
		except:
			colour[sense]=colours.pop()

# col = {'Comedy':'darkblue',
# 'Essays':'lightblue',
# 'Letters':'red',
# 'Narrative':'orange',
# 'Oratory':'maroon',
# 'Philosophy':'green',
# 'Poetry':'gold',
# 'Religion':'chartreuse',
# 'Technical':'salmon',
# 'Tragedy':'purple'}

senses = {}
word_scores = {}
for (sense,genre,century),score in scores.items():
	senses.setdefault(sense,{}).setdefault(genre,{}).setdefault(century,score)
	word=sense[:sense.find(sep)]
	word_scores.setdefault(word,{}).setdefault(century,0)
	word_scores[word][century]+=score
	word_scores.setdefault(sense,{}).setdefault(century,0)
	word_scores[sense][century]+=score
	word_scores.setdefault(word,{}).setdefault(genre,0)
	word_scores[word][genre]+=score
#by genre
print('Generating plots per genre')
genres = {}
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find(sep)]
	genres.setdefault(genre,{}).setdefault(word,{}).setdefault(sense,{}).setdefault(century,score)


#a plot per genre
for genre,words in genres.items():
	colLabels=[str(x) for x in [-8,-7,-6,-5,-4,-3,-2,-1,1,2,3,4,5]]
	Ncols=numpy.arange(len(colLabels))
	fig,ax=plot.subplots()
	ax.set_xlabel('Centuries')
	ax.set_xticks(Ncols)
	ax.set_xticklabels(colLabels)
	ax.set_title(genre)
	ax.set_ylabel('Raw frequency')
	ax.set_xticks(Ncols)
	ax.set_xticklabels(colLabels)
	ax.set_title(genre)
	width = 0.2
	colPositions = []
	if len(words.items()) == 1: colPositions.append(Ncols)
	elif len(words.items()) == 2:
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
	elif len(words.items()) == 3:
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
	for word_idx,(word,senses) in enumerate(words.items()):
		scores_plot_totals = [0 for x in colLabels]
		for sense_idx,(sense,centuries) in enumerate(senses.items()):
			scores_plot = []
			for x in colLabels:
				scores_plot.append(centuries.get(x, 0))
			ax.bar(colPositions[word_idx],scores_plot, width,bottom=scores_plot_totals,color=colour[sense],label='%s: %s'%(word,word_senses[sense]))
			scores_plot_totals=[scores_plot_totals[idx]+x for idx,x in enumerate(scores_plot)]		
	ax.legend(loc='best', fontsize='x-small')
	plot.savefig('%s/plots/by_genre/%s-%s.png'%(dir,word,genre), format='png')

#by time
print('Generating plots per century')
centuries = {}
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find(sep)]
	centuries.setdefault(century,{}).setdefault(word,{}).setdefault(sense,{}).setdefault(genre,score)	

colLabels=[str(x) for x,t in genres.items()]
for century,words in centuries.items():
	Ncols=numpy.arange(len(colLabels))
	fig,ax=plot.subplots()
	ax.set_xticks(Ncols)
	ax.set_xticklabels(colLabels,rotation='vertical')
	ax.set_xlabel('Genres')
	ax.set_ylabel('Raw frequency')
	if int(century)<0:
		century=century[1:]+' BC'
	else:
		century=century+' AD'
	ax.set_title(century)
	width = 0.2
	colPositions = []
	if len(words.items()) == 1: colPositions.append(Ncols)
	elif len(words.items()) == 2:
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
	elif len(words.items()) == 3:
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
	for word_idx,(word,senses) in enumerate(words.items()):
		scores_plot_totals = [0 for x in colLabels]
		for sense_idx,(sense,genres) in enumerate(senses.items()):
			scores_plot = []
			for x in colLabels:
 				scores_plot.append(genres.get(x, 0))
			ax.bar(colPositions[word_idx],scores_plot, width,bottom=scores_plot_totals,color=colour[sense],label='%s: %s'%(word,word_senses[sense]))
			scores_plot_totals=[scores_plot_totals[idx]+x for idx,x in enumerate(scores_plot)]
	ax.legend(loc='best', fontsize='x-small')
	plot.tight_layout()
	plot.savefig('%s/plots/by_century/%s.png'%(dir,century), format='png')
	
#by sense
# print('Generating plots per sense')
# 
# #a plot per sense
# #time variation
# for sense,genres in senses.items():
# 	word=sense[:sense.find(sep)]
# 	colLabels=[str(x) for x in [-7,-6,-5,-4,-3,-2,-1,1,2,3,4,5]]
# 	Ncols=numpy.arange(len(colLabels))
# 	fig,ax=plot.subplots()
# 	ax.set_xlabel('Centuries')
# 	ax.set_ylabel('Frequency (%) of sense per century')
# 	ax.set_xticks(Ncols)
# 	ax.set_xticklabels(colLabels)
# 	ax.set_title('%s: %s'%(sense,word_senses[sense]))
# 	ax.set_ylim([0,107])
# 	width = 0.4
# 	scores_plot_totals = [0 for x in colLabels]
# 	labels = {}
# 	[labels.setdefault(x, [Ncols[idx], 0, '']) for idx,x in enumerate(colLabels)]
# 	for genre,century in genres.items():
# 		scores_plot = []
# 		for idx, x in enumerate(colLabels):
# 			scores_plot.append(century.get(x, 0)*100/word_scores[word].get(str(x),1))
# 			scoresum = scores_plot[idx]+scores_plot_totals[idx]
# 			if scoresum > labels[x][1]:
# 				labels[x][1] = scoresum
# 				labels[x][2] = '%s%%'%round(scoresum,2)
# 		ax.bar(Ncols,scores_plot, width,bottom=scores_plot_totals,color=col[genre],label=genre)
# 		scores_plot_totals=[scores_plot_totals[idx]+x for idx,x in enumerate(scores_plot)]
# 	del scoresum
# 	for x in labels.values():
# 		ax.text(x[0], x[1]+1, x[2], fontsize=8, horizontalalignment='center')
# # calculate totals per column:
# 	err_ranges = []
# 	for i,x in enumerate(colLabels):
# 		try:
# 			err_ranges.append(proportion.proportion_confint(word_scores[sense].get(str(x),0),word_scores[word].get(str(x),0),method='beta'))
# 		except:
# 			err_ranges.append((0,0))
# 	err_lower = [scores_plot[i]-w for i,w in enumerate([x*100 for x,y in err_ranges])]
# 	err_upper = [w-scores_plot[i] for i,w in enumerate([y*100 for x,y in err_ranges])]
# 	ax.errorbar(Ncols, scores_plot, yerr=[err_lower,err_upper], fmt='none', ecolor='black', elinewidth=0.5)
# 	ax.legend(loc='best', fontsize='x-small')
# 	plot.savefig('%s/plots/by_sense/%s-time.png'%(dir,sense), format='png')
# #genre variation	
# 	fig,ax=plot.subplots()
# 	ax.set_xlabel('Centuries')
# 	ax.set_ylabel('Distribution of genres per century (%)')
# 	ax.set_xticks(Ncols)
# 	ax.set_xticklabels(colLabels)
# 	ax.set_title('%s: %s'%(sense,word_senses[sense]))
# 	ax.set_ylim([0,103])
# 	width = 0.4
# 	scores_plot_totals = [0 for x in colLabels]
# 	labels = {}
# 	[labels.setdefault(x, [Ncols[idx], 0, '']) for idx,x in enumerate(colLabels)]
# 	for genre,century in genres.items():
# 		scores_plot = []
# 		for idx, x in enumerate(colLabels):
# 			scores_plot.append(century.get(x, 0)*100/word_scores[sense].get(str(x),1))
# 		ax.bar(Ncols,scores_plot, width,bottom=scores_plot_totals,color=col[genre],label=genre)
# 		scores_plot_totals=[scores_plot_totals[idx]+x for idx,x in enumerate(scores_plot)]
# 	ax.legend(loc='best', fontsize='x-small')
# 	plot.savefig('%s/plots/by_sense/%s-genre.png'%(dir,sense), format='png')

print('Generating a plot per word by century')

genres = set()
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find(sep)]
	genres.add(genre)
	
words = {}
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find(sep)]
	words.setdefault(word,{}).setdefault(sense,{}).setdefault(century,0)
	words[word][sense][century]+=score

w_g_sc = {}
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find(sep)]
	w_g_sc.setdefault(word,{}).setdefault(century,{}).setdefault(genre,0)
	w_g_sc[word][century][genre]+=score

for word,senses in words.items():
	#plot prelims
	colLabels=[str(x) for x in [-8,-7,-6,-5,-4,-3,-2,-1,1,2,3,4,5]]
	Ncols=numpy.arange(len(colLabels))
	width = round(0.66/len(senses),2)
	
	#plot area
	fig,ax=plot.subplots()
	ax.set_xlabel('Centuries')
	ax.set_xticks(Ncols)
	ax.set_xticklabels(colLabels)
	ax.set_title(word)
	ax.set_ylabel('Frequency (%) of sense per century')
	ax.set_ylim([0,107])
	minor_ticks = numpy.arange(-6.5, 13.5, 1)
	ax.set_xticks(minor_ticks, minor=True)
	ax.grid(color='b', linestyle='dotted', linewidth=0.1, axis = 'x', which='minor')
	
	#possible positions of sense-columns
	colPositions = []
	if len(senses.items()) == 1: colPositions.append(Ncols)
	elif len(senses.items()) == 2:
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
	elif len(senses.items()) == 3:
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
	elif len(senses.items()) == 4:
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
	elif len(senses.items()) == 5:
		colPositions.append([x-2*width for x in Ncols])
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
		colPositions.append([x+2*width for x in Ncols])
	elif len(senses.items()) == 6:
		colPositions.append([x-5*width/2 for x in Ncols])
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
		colPositions.append([x+5*width/2 for x in Ncols])
	elif len(senses.items()) == 7:
		colPositions.append([x-3*width for x in Ncols])
		colPositions.append([x-2*width for x in Ncols])
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
		colPositions.append([x+2*width for x in Ncols])
		colPositions.append([x+3*width for x in Ncols])
	elif len(senses.items()) == 8:
		colPositions.append([x-7*width/2 for x in Ncols])
		colPositions.append([x-5*width/2 for x in Ncols])
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
		colPositions.append([x+5*width/2 for x in Ncols])
		colPositions.append([x+7*width/2 for x in Ncols])
	elif len(senses.items()) == 9:
		colPositions.append([x-4*width for x in Ncols])
		colPositions.append([x-3*width for x in Ncols])
		colPositions.append([x-2*width for x in Ncols])
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
		colPositions.append([x+2*width for x in Ncols])
		colPositions.append([x+3*width for x in Ncols])
		colPositions.append([x+4*width for x in Ncols])
	elif len(senses.items()) == 10:
		colPositions.append([x-9*width/2 for x in Ncols])
		colPositions.append([x-7*width/2 for x in Ncols])
		colPositions.append([x-5*width/2 for x in Ncols])
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
		colPositions.append([x+5*width/2 for x in Ncols])
		colPositions.append([x+7*width/2 for x in Ncols])
		colPositions.append([x+9*width/2 for x in Ncols])
	elif len(senses.items()) == 11:
		colPositions.append([x-5*width for x in Ncols])
		colPositions.append([x-4*width for x in Ncols])
		colPositions.append([x-3*width for x in Ncols])
		colPositions.append([x-2*width for x in Ncols])
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
		colPositions.append([x+2*width for x in Ncols])
		colPositions.append([x+3*width for x in Ncols])
		colPositions.append([x+4*width for x in Ncols])
		colPositions.append([x+5*width for x in Ncols])
	elif len(senses.items()) == 12:
		colPositions.append([x-11*width/2 for x in Ncols])
		colPositions.append([x-9*width/2 for x in Ncols])
		colPositions.append([x-7*width/2 for x in Ncols])
		colPositions.append([x-5*width/2 for x in Ncols])
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
		colPositions.append([x+5*width/2 for x in Ncols])
		colPositions.append([x+7*width/2 for x in Ncols])
		colPositions.append([x+9*width/2 for x in Ncols])
		colPositions.append([x+11*width/2 for x in Ncols])
	
	#score labels above bars		
# 	labels = {}
# 	[labels.setdefault(x, [Ncols[idx], 0, '']) for idx,x in enumerate(colLabels)]
	
	#compute scores and bars
	sort_sense = sorted(senses)
	for sense_idx,y in enumerate(sort_sense):
		sense = y
		century = senses[y]
		scores_plot = []
		genres_plot = {}
		for x in colLabels:
			total_score_cen=word_scores[word].get(str(x),1)
			sense_in_cen=century.get(x, 0)*100/total_score_cen
			for genre in genres:
				genres_plot.setdefault(genre,[]).append(w_g_sc.get(word,{}).get(x,{}).get(genre,0)*100/total_score_cen)
# 			labels[x][1] = sense_in_cen
# 			labels[x][2] = '%s%%'%round(sense_in_cen,2)
# 			if sense_in_cen ==  0:
# 				labels[x][2] = ''
			scores_plot.append(sense_in_cen)
		ax.bar(colPositions[sense_idx],scores_plot, width,color=colour[sense],label='%s: %s'%(word,word_senses[sense]))
# 		for x in labels.values():
# 			ax.text(x[0], x[1]+1, x[2], fontsize=8, horizontalalignment='center')

		#error bars
		err_ranges = []
		for i,x in enumerate(colLabels):
			try:
				err_ranges.append(proportion.proportion_confint(century.get(x, 0),word_scores[word].get(str(x),0),method='beta'))
			except:
				err_ranges.append((0,0))
		err_lower = [scores_plot[i]-w for i,w in enumerate([x*100 for x,y in err_ranges])]
		err_upper = [w-scores_plot[i] for i,w in enumerate([y*100 for x,y in err_ranges])]
		ax.errorbar(colPositions[sense_idx], scores_plot, yerr=[err_lower,err_upper], fmt='none', ecolor='black', elinewidth=0.2)
	for genre in genres:
		ax.plot(Ncols, genres_plot[genre],linewidth=0.8,label=genre, ls=':')
	#write plots to files
	chartBox = ax.get_position()
	ax.set_position([chartBox.x0, chartBox.y0, chartBox.width*0.8, chartBox.height])
	ax.legend(loc='upper left', bbox_to_anchor=(1, 1), shadow=False, ncol=1, fontsize='x-small')
	#ax.legend(loc='best', fontsize='x-small')
	#plot.tight_layout()
	plot.savefig('%s/plots/by_sense/by_century/%s.png'%(dir,word), format='png', dpi=500)
	
	
print('Generating a plot per word by genre')

words = {}
for (sense,genre,century),score in scores.items():
	word=sense[:sense.find(sep)]
	words.setdefault(word,{}).setdefault(sense,{}).setdefault(genre,0)
	words[word][sense][genre]+=score
	
for word,senses in words.items():
	#plot prelims
	colLabels=[str(x) for x in genres]
	Ncols=numpy.arange(len(colLabels))
	width = round(0.66/len(senses),2)
	
	#plot area
	fig,ax=plot.subplots()
	ax.set_xlabel('Genres')
	ax.set_xticks(Ncols)
	ax.set_xticklabels(colLabels,rotation='vertical')
	ax.set_title(word)
	ax.set_ylabel('Frequency (%) of sense per genre')
	ax.set_ylim([0,107])
	minor_ticks = numpy.arange(-6.5, 13.5, 1)
	ax.set_xticks(minor_ticks, minor=True)
	ax.grid(color='b', linestyle='dotted', linewidth=0.1, axis = 'x', which='minor')
	
	#possible positions of sense-columns
	colPositions = []
	if len(senses.items()) == 1: colPositions.append(Ncols)
	elif len(senses.items()) == 2:
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
	elif len(senses.items()) == 3:
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
	elif len(senses.items()) == 4:
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
	elif len(senses.items()) == 5:
		colPositions.append([x-2*width for x in Ncols])
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
		colPositions.append([x+2*width for x in Ncols])
	elif len(senses.items()) == 6:
		colPositions.append([x-5*width/2 for x in Ncols])
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
		colPositions.append([x+5*width/2 for x in Ncols])
	elif len(senses.items()) == 7:
		colPositions.append([x-3*width for x in Ncols])
		colPositions.append([x-2*width for x in Ncols])
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
		colPositions.append([x+2*width for x in Ncols])
		colPositions.append([x+3*width for x in Ncols])
	elif len(senses.items()) == 8:
		colPositions.append([x-7*width/2 for x in Ncols])
		colPositions.append([x-5*width/2 for x in Ncols])
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
		colPositions.append([x+5*width/2 for x in Ncols])
		colPositions.append([x+7*width/2 for x in Ncols])
	elif len(senses.items()) == 9:
		colPositions.append([x-4*width for x in Ncols])
		colPositions.append([x-3*width for x in Ncols])
		colPositions.append([x-2*width for x in Ncols])
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
		colPositions.append([x+2*width for x in Ncols])
		colPositions.append([x+3*width for x in Ncols])
		colPositions.append([x+4*width for x in Ncols])
	elif len(senses.items()) == 10:
		colPositions.append([x-9*width/2 for x in Ncols])
		colPositions.append([x-7*width/2 for x in Ncols])
		colPositions.append([x-5*width/2 for x in Ncols])
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
		colPositions.append([x+5*width/2 for x in Ncols])
		colPositions.append([x+7*width/2 for x in Ncols])
		colPositions.append([x+9*width/2 for x in Ncols])
	elif len(senses.items()) == 11:
		colPositions.append([x-5*width for x in Ncols])
		colPositions.append([x-4*width for x in Ncols])
		colPositions.append([x-3*width for x in Ncols])
		colPositions.append([x-2*width for x in Ncols])
		colPositions.append([x-width for x in Ncols])
		colPositions.append(Ncols)
		colPositions.append([x+width for x in Ncols])
		colPositions.append([x+2*width for x in Ncols])
		colPositions.append([x+3*width for x in Ncols])
		colPositions.append([x+4*width for x in Ncols])
		colPositions.append([x+5*width for x in Ncols])
	elif len(senses.items()) == 12:
		colPositions.append([x-11*width/2 for x in Ncols])
		colPositions.append([x-9*width/2 for x in Ncols])
		colPositions.append([x-7*width/2 for x in Ncols])
		colPositions.append([x-5*width/2 for x in Ncols])
		colPositions.append([x-3*width/2 for x in Ncols])
		colPositions.append([x-width/2 for x in Ncols])
		colPositions.append([x+width/2 for x in Ncols])
		colPositions.append([x+3*width/2 for x in Ncols])
		colPositions.append([x+5*width/2 for x in Ncols])
		colPositions.append([x+7*width/2 for x in Ncols])
		colPositions.append([x+9*width/2 for x in Ncols])
		colPositions.append([x+11*width/2 for x in Ncols])
	
	#score labels above bars		
# 	labels = {}
# 	[labels.setdefault(x, [Ncols[idx], 0, '']) for idx,x in enumerate(colLabels)]
	
	#compute scores and bars
	sort_sense = sorted(senses)
	for sense_idx,y in enumerate(sort_sense):
		sense = y
		genre = senses[y]
		scores_plot = []
		for x in colLabels:
			sense_in_gen=genre.get(x, 0)*100/word_scores[word].get(str(x),1)
# 			labels[x][1] = sense_in_gen
# 			labels[x][2] = '%s%%'%round(sense_in_gen,2)
# 			if sense_in_gen ==  0:
# 				labels[x][2] = ''
			scores_plot.append(sense_in_gen)
		ax.bar(colPositions[sense_idx],scores_plot, width,color=colour[sense],label='%s: %s'%(word,word_senses[sense]))
# 		for x in labels.values():
# 			ax.text(x[0], x[1]+1, x[2], fontsize=8, horizontalalignment='center')

		#error bars
		err_ranges = []
		for i,x in enumerate(colLabels):
			try:
				err_ranges.append(proportion.proportion_confint(genre.get(x, 0),word_scores[word].get(str(x),0),method='beta'))
			except:
				err_ranges.append((0,0))
		err_lower = [scores_plot[i]-w for i,w in enumerate([x*100 for x,y in err_ranges])]
		err_upper = [w-scores_plot[i] for i,w in enumerate([y*100 for x,y in err_ranges])]
		ax.errorbar(colPositions[sense_idx], scores_plot, yerr=[err_lower,err_upper], fmt='none', ecolor='black', elinewidth=0.2)

	#write plots to files
	ax.legend(loc='best', fontsize='x-small')
	plot.tight_layout()
	plot.savefig('%s/plots/by_sense/by_genre/%s.png'%(dir,word), format='png', dpi=500)