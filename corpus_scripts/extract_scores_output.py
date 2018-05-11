import os
import re
import configparser
import numpy as np
import pandas as pd
import sys
from grkLemmata import greekLemmata
from lxml import etree as document
from openpyxl import load_workbook
from scipy import stats
import pickle

def century(year): #change into 0 – 11 (boh)
	year = int(year)
	if year >= -800 and year <= -700:
		return 0 #-7
	elif year > -700 and year <= -600:
		return 1 #-7
	elif year > -600 and year <= -500:
		return 2 #-6
	elif year > -500 and year <= -400:
		return 3 #-5
	elif year > -400 and year <= -300:
		return 4 #-4
	elif year > -300 and year <= -200:
		return 5 #-3
	elif year > -200 and year <= -100:
		return 6 #-2
	elif year > -100 and year <= -1:
		return 7 #-1
	elif year >= 0 and year < 100:
		return 8 #1
	elif year >= 100 and year < 200:
		return 9 #2
	elif year >= 200 and year < 300:
		return 10 #3
	elif year >= 300 and year < 400:
		return 11 #4
	elif year >= 400 and year < 500:
		return 12 #5

os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))
config = configparser.ConfigParser()
config.read('config.ini')
dir = config['paths']['output']
file_list = config['paths']['file_list']
input = config['paths']['model_output']
o_dir = '%s/%s'%(dir,'var_change')

output_file = '%s/50_words_variation_data_auto_K_micro_and_macro_S.txt'%o_dir
if os.path.isfile(output_file): os.remove(output_file)
output = open(output_file, 'a')

carry_on = True
scores = {}

subgenres = set()
genres = set()
summary_g = set()
summary_sg = set()

for currDir, dirs, files in os.walk(input):
	for word in dirs:
		scores.setdefault(word, {})

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]
af = config['paths']['final_corpus']

if 'regenerate' in sys.argv:
	if os.path.isfile("%s/50_w_data.p"%o_dir):
		os.remove("%s/50_w_data.p"%o_dir)
		
if os.path.isfile("%s/50_w_data.p"%o_dir):
	scores = pickle.load(open("%s/50_w_data.p"%o_dir, "rb" ))
else:
	for record in files:
		subgenre = record[h['Subgenre']].value
		subgenres.add(subgenre)
		genre = record[h['Genre']].value
		genres.add(genre)
		file = '%s/%s'%(af,record[h['Tokenized file']].value)
		cent = century(record[h['Date']].value)
		sys.stdout.write("\r\033[KProcessing %s"%record[h['Tokenized file']].value)
		sys.stdout.flush()
		curr_text = document.parse(file)
		for word,score in scores.items():
			#print(word, len(curr_text.xpath('//word/lemma[@id="%s"]'%word)),subgenre,cent)
			ref=(cent,subgenre)
			scores[word].setdefault('subgenre',{}).setdefault(ref,0)
			scores[word]['subgenre'][ref] += len(curr_text.xpath('//word/lemma[@id="%s"]'%word))
			ref=(cent,genre)
			scores[word].setdefault('genre',{}).setdefault(ref,0)
			scores[word]['genre'][ref] += len(curr_text.xpath('//word/lemma[@id="%s"]'%word))
			scores[word].setdefault('century',{}).setdefault(cent,0)
			scores[word]['century'][cent] += len(curr_text.xpath('//word/lemma[@id="%s"]'%word))
	pickle.dump(scores, open("%s/50_w_data.p"%o_dir, "wb" ))

for currDir, dirs, files in os.walk(input):
	for word in dirs:
		file = open('%s/%s/output.dat'%(currDir,word), 'r')
		carry_on = True
		for line in file:
			if carry_on == True:
				try:
					line.split()[1][0:1]
				except IndexError:
					continue
				else:					
					if line.split()[1][0:1] == 'T':
						value = (int(re.search('K=(.*?):',line.split()[1]).group(1)),int(re.search('T=(.*?),',line.split()[1]).group(1)))
						score = float(line.split()[0])
						scores[word].setdefault('results',{}).setdefault(value, score)
						del value
					elif line.split()[1] == 'per' and line.split()[2] == 'time':
						carry_on = False

for word,score in scores.items():
	lines_g = {}
	lines_sg = {}
	lines_c = {}
	try:
		results = score['results']
	except:
		continue
	else:
		titleStr= '%s (%s)'%(greekLemmata[word]['lemma'], word)
		output.write(100*'#'+'\n#   '+titleStr+(95-len(titleStr))*' '+'#\n'+100*'#'+'\n')
		print(word, greekLemmata[word]['lemma'])
		genres = score['genre']
		subgenres = score['subgenre']
		centuries = score['century']
		for (c,g),sc in genres.items():
			for d in range(13):
				lines_g.setdefault((g,d),0)
			lines_g[(g,c)]=sc
		for (c,g),sc in subgenres.items():
			for d in range(13):
				lines_sg.setdefault((g,d),0)
			lines_sg[(g,c)]=sc
		for c,sc in centuries.items():
			for d in range(13):
				lines_c.setdefault(d,1)
			if sc > 0: lines_c[c]=sc
		lines_r={}	
		for (k,t),prob in results.items():
			lines_r.setdefault('#K%s'%k, {}).setdefault(t,prob)
			for (g,c),sc in lines_g.items():
				lines_r.setdefault('G:%s'%g, {}).setdefault(c,sc)
			for (g,c),sc in lines_sg.items():
				lines_r.setdefault('SG:%s'%g, {}).setdefault(c,sc)
		lines_r.setdefault('T',lines_c)
		table=pd.DataFrame(data=lines_r)
		if table.isnull().values.any() == True:
			continue
	#	with pd.option_context('expand_frame_repr', False):
	#		print(table)
		Ks=[col for col in table if col.startswith('#')]
		Gs=[col for col in table if col.startswith('G')]
		Sgs=[col for col in table if col.startswith('SG')]
		for k in Ks:
			#correlation with each Gs
			output.write('\n%s:%s ~ correlation with genres\n'%(greekLemmata[word]['lemma'],k[1:]))
			corr_with = []
			corr_ranking = []
			for g in Gs:
				if sum(table[g]) == 0:
					r=(np.nan,np.nan)
				else:
					r=stats.spearmanr(table[k], table[g].div(table['T']))
				corr_ranking.append(('%s: %s ~ %s'%(greekLemmata[word]['lemma'],k[1:],g),r[0],r[1]))
			ranking=pd.DataFrame([x[1:] for x in corr_ranking], columns=['Spearman\'s ρ','p'],index=[x[0] for x in corr_ranking])
			ranking=ranking.reindex(ranking[['Spearman\'s ρ']].abs().sort_values(by=['Spearman\'s ρ'], ascending=False).index)
			print(ranking,'\n')
			output.write('\n'+ranking.to_string()+'\n\n')
			for x in range(len(ranking.loc[ranking['p']<=0.05])):
				cool_row=ranking.loc[ranking['p']<=0.05].iloc[x]
				corr_with.append('%s (Spearman\'s ρ = %s, p = %s)'%(cool_row.name.split('~ ')[1],round(cool_row[0],3),round(cool_row[1],3)))
				del cool_row
			list_corr = '\n\t\t'.join(corr_with)
			if len(list_corr) == 0:
				list_corr = 'no particular genre'
			else:
				list_corr='\n\t\t%s'%list_corr
			summary_g.add('%s: %s is correlated with %s'%(greekLemmata[word]['lemma'],k,list_corr))
			del list_corr
			
			#correlation with each Sgs
			output.write('\n%s:%s ~ correlation with subgenres\n'%(greekLemmata[word]['lemma'],k[1:]))
			corr_with=[]
			corr_ranking = []
			for g in Sgs:
				if sum(table[g]) == 0:
					r=(np.nan,np.nan)
				else:
					r=stats.spearmanr(table[k], table[g].div(table['T']))
				corr_ranking.append(('%s: %s ~ %s'%(greekLemmata[word]['lemma'],k[1:],g),r[0],r[1]))
			ranking=pd.DataFrame([x[1:] for x in corr_ranking], columns=['Spearman\'s ρ','p'],index=[x[0] for x in corr_ranking])
			ranking=ranking.reindex(ranking[['Spearman\'s ρ']].abs().sort_values(by=['Spearman\'s ρ'], ascending=False).index)
			print(ranking,'\n')
			output.write('\n'+ranking.to_string()+'\n\n')
			for x in range(len(ranking.loc[ranking['p']<=0.05])):
				cool_row=ranking.loc[ranking['p']<=0.05].iloc[x]
				corr_with.append('%s (Spearman\'s ρ = %s, p = %s)'%(cool_row.name.split('~ ')[1],round(cool_row[0],3),round(cool_row[1],3)))
				del cool_row
			list_corr = '\n\t\t'.join(corr_with)
			if len(list_corr) == 0:
				list_corr = 'no particular subgenre'
			else:
				list_corr='\n\t\t%s'%list_corr
			summary_sg.add('%s: %s is correlated with %s'%(greekLemmata[word]['lemma'],k,list_corr))
			del list_corr
		print('\n\n\n')
		
print('\n\n### SUMMARY: CORRELATIONS WITH GENRES ###\n\n')		
print('\n'.join(sorted(summary_g)))
output.write('\n\n### SUMMARY: CORRELATIONS WITH GENRES ###\n\n')
output.write('\n'.join(sorted(summary_g)))
del summary_g

print('\n\n### SUMMARY: CORRELATIONS WITH SUBGENRES ###\n\n')
print('\n'.join(sorted(summary_sg)))
output.write('\n\n### SUMMARY: CORRELATIONS WITH SUBGENRES ###\n\n')
output.write('\n'.join(sorted(summary_sg)))
del summary_sg
output.close()