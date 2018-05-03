import sys
from lxml import etree as document
from openpyxl import load_workbook
import configparser
import os
from stop_cltk import STOPS_LIST_ID
import re
from beta2utf import convertBeta

os.chdir(os.path.dirname(os.path.realpath(__file__)))
os.system("clear && printf '\e[3J'")
config = configparser.ConfigParser()
config.read('config.ini')
annotated = config['paths']['annotated']
resources = config['paths']['file_list']
dir = config['paths']['output']
TT = config['paths']['treetagger_output']

wb = load_workbook('%s/word_senses.xlsx'%resources)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h_sense = {cell.value : n for n, cell in enumerate(headers[0])}
data = ws['A2:E23']

wb2 = load_workbook('%s/file_list.xlsx'%resources)
ws2 = wb2.active
headers = ws2[config['excel_range']['headers']]
h_file = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws2[config['excel_range']['range']]

fields = ['date', 'genre', 'author', 'work', 'tlg author', 'tlg work', 'sentence location', 'sentence id', 'sentence ids', 'sentence original', 'target id', 'sense id', 'disambiguation','subgenre']

target_words = {x for x in sys.argv[1:]}
if len(target_words) == 0:
	target_words = {x[h_sense['TERM ID']].value for x in data}
	
tw_files = {}
for tw in target_words:
	output_file_name='%s/senses_%s.txt'%(dir,tw)
	if os.path.isfile(output_file_name): os.remove(output_file_name)	
	tw_files[tw]=open(output_file_name, 'a')
	tw_files[tw].write('\t'.join(fields))
	
for idx,record in enumerate(files):
	file = '%s/%s'%(annotated,record[h_file['Tokenized file']].value)
	date = str(record[h_file['Date']].value)
	#century = str(record[h_file['Century']].value)
	genre = str(record[h_file['Genre']].value)
	subgenre = str(record[h_file['Subgenre']].value)
	author = str(record[h_file['Author']].value)
	tlg_work = str(record[h_file['TLG ID']].value)
	tlg_author = str(record[h_file['TLG Author']].value)
	work = str(record[h_file['Work']].value)
	sys.stdout.write("\r\033[KChecking %s"%os.path.basename(file))
	sys.stdout.flush()
	curr_text = open(file,'r').read()
	if curr_text.find('sense') == -1: continue
	del curr_text
	
	curr_text = document.parse(file)
	sys.stdout.write("\r\033[KExtracting data from %s"%os.path.basename(file))
	sys.stdout.flush()
	TT_doc = open('%s/%s'%(TT,record[h_file['Tokenized file']].value.replace('xml','txt')), 'r')
	TT_lines=[x for x in TT_doc]
	finalTxt = ''
	nodes = curr_text.xpath('//sentence|//word|//punct')
	word_count = 0
	for node in nodes:
		if node.tag == 'sentence':
			finalTxt += '\n'
			finalTxt += '%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s\t'%(date,genre,author,work,tlg_author,tlg_work,node.get('location'),node.get('id'),subgenre)			
		elif node.tag == 'word':
			lemma_count = len(node.xpath('./lemma'))
			if lemma_count == 1:
				finalTxt+='*%s*'%node.xpath('lemma/@id')[0]
			elif lemma_count > 1:
				TT_POS = TT_lines[word_count].split('\t')[1].strip()
				try:
					finalTxt+='*%s*'%node.xpath('lemma[@POS="%s"]/@id'%TT_POS)[0]
				except:
					if TT_POS == 'proper':
						TT_POS='noun'
					try:
						finalTxt+='*%s*'%node.xpath('lemma[@POS="%s"]/@id'%TT_POS)[0]
					except:
						TT_POS='adjective'
						try:
							finalTxt+='*%s*'%node.xpath('lemma[@POS="%s"]/@id'%TT_POS)[0]
						except:
							finalTxt+='*%s*'%node.xpath('lemma/@id')[0]	
			word_count+=1				
		elif node.tag == 'punct':
			word_count+=1
	del TT_doc, TT_lines, nodes, lemma_count, TT_POS
	
	for stop in STOPS_LIST_ID:
		finalTxt = re.sub('\*%s\*'%stop, '', finalTxt)
	finalTxt = re.sub('.*?\t\n','',finalTxt)
	
	for tw,fname in tw_files.items():
		ttw = ''
		allInstances = re.findall('.*?\*'+tw+'\*.*?\n', finalTxt)
		if len(allInstances) > 0: fname.write('\n')
		for instance in allInstances:
			sentence_id = instance.split('\t')[7]
			sense = curr_text.xpath('//sentence[@id="%s"]/word/lemma[@id="%s"]'%(sentence_id, tw))[0].get('sense')
			notes = curr_text.xpath('//sentence[@id="%s"]/word/lemma[@id="%s"]'%(sentence_id, tw))[0].get('notes')
			form = ' '.join(convertBeta(x.get('form')) for x in curr_text.xpath('//sentence[@id="%s"]/word'%sentence_id))
			instance = instance.replace('*\n', '*\t%s\t%s\t%s\t%s\n'%(form,tw,sense,notes))
			instance = instance.replace('**', ' ').replace('*', '')
			ttw += instance	
		fname.write(ttw.strip())
	del curr_text, finalTxt, ttw
for fname in tw_files.values():
	fname.close()
sys.stdout.write("\r\033[KAll done!\n")
sys.stdout.flush()