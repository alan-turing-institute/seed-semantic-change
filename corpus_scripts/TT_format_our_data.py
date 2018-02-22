import os
import configparser
from openpyxl import load_workbook
from lxml import etree as document
import sys

os.chdir(os.path.dirname(os.path.realpath(__file__)))
os.system("clear && printf '\e[3J'")
config = configparser.ConfigParser()
config.read('config.ini')
treetagger = config['paths']['treetagger']
file_list = config['paths']['file_list']
af = config['paths']['annotated']
tf = config['paths']['tokenized']
for_tt = config['paths']['for_treetagger']

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]

#Prepare our data
print('Formatting texts')
for record in files:
	author = record[h['Author']].value
	work = record[h['Work']].value
	genre = record[h['Genre']].value
	path = '%s/%s'%(tf,record[h['Tokenized file']].value)
	open('%s/%s'%(for_tt,record[h['Tokenized file']].value.replace('xml','txt')), 'w')
	output_singlefile=open('%s/%s'%(for_tt,record[h['Tokenized file']].value.replace('xml','txt')), 'a')
	print('Formatting %s: %s'%(author, work))
	parse = document.parse(path)
	tokens = parse.xpath('//word|//punct')
	for token in tokens:
		if token.tag == 'word':
			output_singlefile.write('%s\n'%token.get('form'))
		elif token.tag == 'punct':
			output_singlefile.write('%s\n'%token.get('mark'))
