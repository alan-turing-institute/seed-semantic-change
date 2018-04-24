import re
import os
import sys
from openpyxl import load_workbook
import configparser
from stop_cltk import STOPS_LIST_ID
from grkLemmata import greekLemmata
from lxml import etree as document
os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))

config = configparser.ConfigParser()
config.read('config.ini')

file_list = config['paths']['file_list']
af = config['paths']['final_corpus']
dir = config['paths']['output']
filterStopWords=config['params']['filter_stop_words']

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]

fullText=open('%s/full_corpus_forms.txt'%dir, 'w')
fullText_ids=open('%s/full_corpus_ids.txt'%dir, 'w')
finalString = ''
finalString_ids = ''

for idx,record in enumerate(files):
	file = '%s/%s'%(af,record[h['Tokenized file']].value)
	wy = str(record[h['Date']].value)
	sys.stdout.write("\r\033[KProcessing %s"%record[h['Tokenized file']].value)
	sys.stdout.flush()
	finalTxt = open(file, 'r').read().replace("\n", "").replace(" ", "")
	finalTxt = re.sub('.*?<body>(.*?)</body>.*', r'\1', finalTxt)
	finalTxt = re.sub('<punct.*?/>', '', finalTxt)
	finalTxt = re.sub('<sentenceid=".*?"location=".*?"/?>', '['+wy+']\t', finalTxt)
	finalTxt = re.sub('<word.*?lemmaid="(.*?)".*?</word>', r'*\1*', finalTxt)
	finalTxt = re.sub('</sentence>', '\n', finalTxt)
	if filterStopWords == "True":
		for stop in STOPS_LIST_ID:
			finalTxt = re.sub('\*%s\*'%stop, '', finalTxt)
		finalTxt = re.sub('.*?\t\n','',finalTxt)
	finalTxt = finalTxt.replace('**', ' ').replace('*', '')
	finalTxt = re.sub('^-?\d+\n','',finalTxt)
	finalTxt_ids = finalTxt
	conv_text = []
	conv_text_l = []
	for x in finalTxt.split():	
		if re.search('\[.*?\]', x) != None:
			conv_text.append("\n"+x)
			conv_text_l.append("\n"+x)
		else:
			conv_text.append(x)
			conv_text_l.append(greekLemmata[x]['lemma'])
	finalTxt = ' '.join(conv_text_l).replace('] ', ']\t')
	finalTxt = finalTxt.replace('[', '').replace(']', '')
	finalTxt_ids = ' '.join(conv_text).replace('] ', ']\t')
	finalTxt_ids = finalTxt_ids.replace('[', '').replace(']', '')
	finalString += finalTxt
	finalString_ids += finalTxt_ids
fullText.write(finalString.strip())
fullText.close()
fullText_ids.write(finalString_ids.strip())
fullText_ids.close()
print('\n###########\n#All done!#\n###########\n')