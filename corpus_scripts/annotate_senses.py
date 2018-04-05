import os
import sys
import re
import configparser
from lxml import etree
from openpyxl import load_workbook
from grkLemmata import greekLemmata
from beta2utf import convertBeta
import operator
import textwrap

os.chdir(os.path.dirname(os.path.realpath(__file__)))
os.system("printf '\e[0m'")
os.system("clear && printf '\e[3J'")
config = configparser.ConfigParser()
config.read('config.ini')
annotated = config['paths']['annotated']
resources = config['paths']['file_list']
window = int(config['params']['window'])

wb = load_workbook('%s/word_senses.xlsx'%resources)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h_sense = {cell.value : n for n, cell in enumerate(headers[0])}
data = ws['A2:E23']

wb2 = load_workbook('%s/file_list.xlsx'%resources)
ws2 = wb2.active
headers = ws2[config['excel_range']['headers']]
h_file = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws2[config['excel_range']['range']]

def sense_in_file():
	for file in files:
		current_file = '%s/%s'%(annotated,file[h_file['Tokenized file']].value)
		parse = etree.parse(current_file)
		target_words_hits = parse.xpath('//word[lemma[@id="%s"][not(@sense)]]'%word_id)
		sys.stdout.write('\r\033[KSearching %s...'%os.path.basename(current_file))
		sys.stdout.flush()
		if len(target_words_hits)>0 :
			sentences = []
			author = parse.xpath('//author')[0].text
			title = parse.xpath('//title')[0].text
			sys.stdout.write('\r\033[K%s - %s (%s hits)\n'%(author, title, len(target_words_hits)))
			sys.stdout.flush()
			for idx,hit in enumerate(target_words_hits):
				sentence=hit.getparent()
				target_word_id = hit.get('id')
				w_i_s = sentence.xpath('./word')
				location = sentence.get('location')
				sentence_form = [convertBeta(x.get('form')) for x in w_i_s]
				sentence_form[int(target_word_id)-1]='\033[1;31;47m\u00a0%s\u00a0\033[0m'%sentence_form[int(target_word_id)-1]
				sentence_form = ' '.join(sentence_form)
				idxStr = '%s/%s'%(idx+1,len(target_words_hits))
				spaces = 12-len(location)-len(idxStr)
				print('%s%s[%s] %s'%(idxStr,spaces*' ', location, textwrap.wrap(sentence_form, width = 100)[0]))
				try:
					print(textwrap.indent('\n'.join(textwrap.wrap(sentence_form, width = 100)[1:]), 15*' '))
				except:
					continue
				sense_input = [x+1 for x, values in enumerate(senses)]
				sense_ids = [x for x in senses.keys()]
				sense_glosses = [x for x in senses.values()]
				prompt_lines = []
				for idx,x in enumerate(sense_input):
					prompt_lines.append('%s: %s'%(x, sense_glosses[idx]))
				try:
					while True:
						i_s = input('Enter sense:\n\t%s\n\t\033[31;47;1m'%'\n\t'.join(prompt_lines))
						os.system("printf '\e[0m'")
						try:
							if int(i_s) <= len(senses) and int(i_s) > 0:
								break
						except:
							continue
				except KeyboardInterrupt:
					os.system("printf '\e[0m'")
				print('Input sense: \033[45;1m %s \033[0m\033[K'%sense_glosses[int(i_s)-1])
				while True:
					notes = input('How did you identify the sense? (1 - collocates ; 2 - background knowledge ; 3 - world knowledge ; 4 - logic (other senses would be absurd) ; 5 - genre ; 6 - register ; 7 - error (lemmatization, tokenization, spelling)) ')
					if int(notes) in range(1,8):
						break
				print()
				hit.xpath('./lemma[@id="%s"]'%word_id)[0].set('notes', notes)
				hit.xpath('./lemma[@id="%s"]'%word_id)[0].set('sense', sense_ids[int(i_s)-1])
				parse.write(current_file, xml_declaration = True, encoding='UTF-8', pretty_print=True)

#extract target word data
word_senses = {}
for record in data:
	word_id = record[h_sense['TERM ID']].value
	word = record[h_sense['TERM']].value
	sense_id = record[h_sense['SENSE LSJ']].value
	sense = record[h_sense['SENSE']].value
	word_senses.setdefault(word_id, {}).setdefault(sense_id, sense)

for word_id, senses in word_senses.items():
	senses['w'] = 'wrong'
	word = greekLemmata[word_id]['lemma']
	print('Target word: %s'%word)
	sense_in_file()