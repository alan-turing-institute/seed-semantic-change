import re
import os
import sys
from openpyxl import load_workbook
import configparser
from stop_cltk import STOPS_LIST_ID
from grkLemmata import greekLemmata
from lxml import etree as document

def readCSV(file):
	wlf = open(file, 'r').read().split('\n')
	read_word_list_file = []
	for row in wlf:
		read_word_list_file.append(row.split(','))
	return read_word_list_file

os.system("clear && printf '\e[3J'")
os.chdir(os.path.dirname(os.path.realpath(__file__)))

config = configparser.ConfigParser()
config.read('config.ini')

file_list = config['paths']['file_list']
af = config['paths']['annotated']
lf = config['paths']['output']
wrl = config['paths']['word_list']
TT = config['paths']['treetagger_output']
filterStopWords=config['params']['filter_stop_words']

wb = load_workbook('%s/file_list.xlsx'%file_list)
ws = wb.active
headers = ws[config['excel_range']['headers']]
h = {cell.value : n for n, cell in enumerate(headers[0])}
files = ws[config['excel_range']['range']]

#prompt word list
if 'no_wordlist' in sys.argv:
	word_list_toggle = False
else:
	while True:
		a = input('Filter sentences using word list? [y/n] ')
		if re.search('[yYnN]', a) != None:
			break
	if re.search('[yY]', a) != None:
		word_list_toggle = True
		fileList = ''
		wordListDict = {}
		w = 0
		for file in os.listdir(wrl):
			if file.endswith(".csv"):
				w+=1
				print(w,':',file)
				wordListDict[w]=file
		wordList = int(input('Insert wordlist file number: '))
		while True :
			try:
				word_list_file=readCSV("%s/%s"%(wrl,wordListDict[wordList]))
			except:
				print('Wrong input')
				wordList = int(input('Insert wordlist file number: '))		
			else:
				break
		word_list = []
		for row in word_list_file:
			word_list.append(row[1])
		words_to_match =''
		for word in word_list:
			words_to_match += '%s|'%word
		words_to_match='\*(?:%s)\*'%words_to_match[:-1]
		search='.*?'+words_to_match+'.*?\n'
	else:
		word_list_toggle = False

#prompt filter by field
if 'no_field' in sys.argv:
	metadata_toggle = False
else:
	while True:
		a = input('Filter sentences by metadata value? [y/n] ')
		if re.search('[yYnN]', a) != None:
			break
	if re.search('[yY]', a) != None:
		metadata_toggle = True
		fields = ''	
		field = ''
		field_value = ''
		for x in h.items():
			fields += (' = '.join([str(x[1]), x[0]]))
			fields += '; '
		fields = fields[:-2]
		field = int(input('Insert field number (%s): '%fields))
		field_value = str(input('Insert field value to filter: '))
	else:
		metadata_toggle = False
	
#prompt TreeTagger disambiguation:
if 'TT' in sys.argv:
	toggle_TT = True
else:
	while True:
		a = input('Use TreeTagger disambiguation of multiple lemmata? [y/n] ')
		if re.search('[yYnN]', a) != None:
			break
	if re.search('[yY]', a) != None:
		toggle_TT = True
	else:
		toggle_TT = False

#prompt words by ID or form
if 'id' in sys.argv:
	toggle_id = True
else:
	while True:
		a = input('Words displayed as IDs or Greek forms? [i/f] ')
		if re.search('[iIfF]', a) != None:
			break
	if re.search('[iI]', a) != None:
		toggle_id = True
	else:
		toggle_id = False

#display sentence location and id?
if 'no_location' in sys.argv:
	toggle_sentence_id = False
else:
	while True:
		a = input('Display sentence ID and location? [y/n] ')
		if re.search('[yYnN]', a) != None:
			break
	if re.search('[yY]', a) != None:
		toggle_sentence_id = True
	else:
		toggle_sentence_id = False
	
#display work id
if 'no_work_id' in sys.argv:
	toggle_work_id = False
else:
	while True:
		a = input('Display Work ID? [y/n] ')
		if re.search('[yYnN]', a) != None:
			break
	if re.search('[yY]', a) != None:
		toggle_work_id = True
	else:
		toggle_work_id = False
	
dir = lf
output_file_name = input('Output file name: ')
fullText=open('%s/%s'%(dir, output_file_name), 'a')
	
for idx,record in enumerate(files):
	file = '%s/%s'%(af,record[h['Tokenized file']].value)
	wy = str(record[h['Date']].value)
	
	if toggle_work_id == True:
		wid = '[work_id:%s.%s]'%(record[h['TLG Author']].value,record[h['TLG ID']].value)
	else:
		wid = ''
		
	if metadata_toggle == True:
		value = str(record[field].value).lower()
		if value == field_value.lower():
			green_light = True
		else:
			green_light = False
	else:		
		green_light = True
	if green_light == False:
		continue
		
	print('Processing',record[h['Tokenized file']].value)	
	if toggle_TT == False:
		finalTxt = open(file, 'r').read().replace("\n", "").replace(" ", "")
		finalTxt = re.sub('.*?<body>(.*?)</body>.*', r'\1', finalTxt)
		finalTxt = re.sub('<punct.*?/>', '', finalTxt)
	
		if toggle_sentence_id == True:
			finalTxt = re.sub('<sentenceid="(.*?)"location="(.*?)">', '['+wy+']'+wid+r'[sentence_id:\1][sentence_loc:\2]\t', finalTxt)
		else:
			finalTxt = re.sub('<sentenceid="(.*?)"location="(.*?)">', '['+wy+']'+wid+'\t', finalTxt)
	
		finalTxt = re.sub('<word.*?lemmaid="(.*?)".*?</word>', r'*\1*', finalTxt)
		finalTxt = re.sub('</sentence>', '\n', finalTxt)
	else:
		#parse as xml
		finalTxt = ''
		curr_text = document.parse(file)
		TT_doc = open('%s/%s'%(TT,record[h['Tokenized file']].value.replace('xml','txt')), 'r')
		TT_lines=[x for x in TT_doc]
		nodes = curr_text.xpath('//sentence|//word|//punct')
		word_count = 0
		for node in nodes:
			if node.tag == 'sentence':
				finalTxt += '\n'
				if toggle_sentence_id == True:
					finalTxt += '[%s]%s[sentence_id:%s][sentence_loc:%s]\t'%(wy,wid,node.get('id'),node.get('location'))
				else:
					finalTxt += '[%s]%s\t'%(wy,wid)
			elif node.tag == 'word':
				lemma_count = len(node.xpath('./lemma'))
				if lemma_count == 1:
					finalTxt+='*%s*'%node.xpath('lemma/@id')[0]
				elif lemma_count > 1:
					TT_POS = TT_lines[word_count].split('\t')[1].strip()
					try:
						finalTxt+='*%s*'%node.xpath('lemma[@POS="%s"]/@id'%TT_POS)[0]
					except:
						if TT_POS == 'proper':
							TT_POS='noun'
						try:
							finalTxt+='*%s*'%node.xpath('lemma[@POS="%s"]/@id'%TT_POS)[0]
						except:
							TT_POS='adjective'
							try:
								finalTxt+='*%s*'%node.xpath('lemma[@POS="%s"]/@id'%TT_POS)[0]
							except:
								finalTxt+='*%s*'%node.xpath('lemma/@id')[0]	
				word_count+=1				
			elif node.tag == 'punct':
				word_count+=1
		del curr_text, TT_doc, TT_lines, nodes, lemma_count, TT_POS

	if filterStopWords == "True":
		for stop in STOPS_LIST_ID:
			finalTxt = re.sub('\*%s\*'%stop, '', finalTxt)
		finalTxt = re.sub('.*?\t\n','',finalTxt)
				
	if word_list_toggle == True:
		allInstances = re.findall(search, finalTxt)
		finalTxt = ''.join(allInstances)

	finalTxt = finalTxt.replace('**', ' ').replace('*', '')
	
	if toggle_id == False:
		conv_text = []
		for x in finalTxt.split():	
			if re.search('\[.*?\]', x) != None:
				conv_text.append("\n"+x)
			else:
				conv_text.append(greekLemmata[x]['lemma'])
		finalTxt = ' '.join(conv_text).replace('] ', ']\t')
		
	finalTxt = finalTxt.replace('[', '').replace(']', '')
	fullText.write(finalTxt.strip())
fullText.close()
print('\n###########\n#All done!#\n###########\n')